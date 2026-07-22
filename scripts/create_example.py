"""Build the sample workbook used by the examples in README.md.

Writes ``examples/example.xlsx`` — a small, plausible café report spread over
four sheets, carrying data in every shape the EDJAS documentation describes:
scalars, a vector, two-column name/value ranges, a header-topped table and a
column-oriented (field-per-row) block.

The companion spec lives in ``examples/example.toml``. Together they let every
construct in the README be demonstrated against real data:

    edjas examples/example.xlsx examples/example.toml

Run from the project root:  ``python scripts/create_example.py``
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

EXAMPLES = Path(__file__).resolve().parent.parent / "examples"
XLSX = EXAMPLES / "example.xlsx"

HOURS = [
    ("Monday", "07:00-20:00"),
    ("Tuesday", "07:00-20:00"),
    ("Wednesday", "07:00-20:00"),
    ("Thursday", "07:00-20:00"),
    ("Friday", "07:00-22:00"),
    ("Saturday", "09:00-22:00"),
    ("Sunday", "Closed"),
]
PRICES = [("Tea", 3.25), ("Coffee", 4.0), ("Bacon roll", 8.25)]
COVERS = [("Mon", "128"), ("Tue", "143"), ("Wed", "97")]  # numbers held as text
CODES = [("GF", "Gluten-free"), ("VG", "Vegan")]
TAGS = ["Vegan", "Gluten-free", "Dairy-free"]
SALES = [
    ("Region", "Q1", "Q2"),
    ("North", 1200, 1350),
    ("South", 980, 1010),
    ("East", 1440, 1390),
]
STAFF = [  # column-oriented: one field per row
    ("name", "Ada", "Grace"),
    ("role", "Barista", "Chef"),
]


def build():
    wb = Workbook()

    # --- Summary: scalars (this is the active sheet) ----------------------
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Riverside Cafe - quarterly summary"
    summary["A2"], summary["B2"] = "Title", "Riverside Cafe"
    summary["A3"], summary["B3"] = "Period ending", datetime(2026, 3, 31)
    summary["A4"], summary["B4"] = "Average spend", 8.7451

    # --- Data: objects and a vector ---------------------------------------
    data = wb.create_sheet("Data")
    for i, (day, opening) in enumerate(HOURS):           # A1:B7
        data.cell(row=1 + i, column=1, value=day)
        data.cell(row=1 + i, column=2, value=opening)
    for i, (item, price) in enumerate(PRICES):           # D1:E3
        data.cell(row=1 + i, column=4, value=item)
        data.cell(row=1 + i, column=5, value=price)
    for i, (day, count) in enumerate(COVERS):            # G1:H3
        data.cell(row=1 + i, column=7, value=day)
        data.cell(row=1 + i, column=8, value=count)
    for i, (code, name) in enumerate(CODES):             # J1:K2
        data.cell(row=1 + i, column=10, value=code)
        data.cell(row=1 + i, column=11, value=name)
    for i, tag in enumerate(TAGS):                       # M1:M3
        data.cell(row=1 + i, column=13, value=tag)

    # --- Sales: a header-topped table --------------------------------------
    sales = wb.create_sheet("Sales")
    for r, row in enumerate(SALES, start=1):             # A1:C4
        for c, value in enumerate(row, start=1):
            sales.cell(row=r, column=c, value=value)

    # --- Staff: field-per-row, needs transposing ---------------------------
    staff = wb.create_sheet("Staff")
    for r, row in enumerate(STAFF, start=1):             # A1:C2
        for c, value in enumerate(row, start=1):
            staff.cell(row=r, column=c, value=value)

    for name, ref in [
        ("Title", "Summary!$B$2"),
        ("PeriodEnd", "Summary!$B$3"),
        ("AvgSpend", "Summary!$B$4"),
        ("Hours", "Data!$A$1:$B$7"),
        ("Prices", "Data!$D$1:$E$3"),
        ("Covers", "Data!$G$1:$H$3"),
        ("Codes", "Data!$J$1:$K$2"),
        ("Tags", "Data!$M$1:$M$3"),
        ("Sales", "Sales!$A$1:$C$4"),
        ("Staff", "Staff!$A$1:$C$2"),
    ]:
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    EXAMPLES.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    print(f"Wrote workbook: {XLSX}")


if __name__ == "__main__":
    build()
