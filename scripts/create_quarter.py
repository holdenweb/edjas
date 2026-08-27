"""Build the filming workbook used by the promotional animation.

Writes ``video/quarter.xlsx`` — a deliberately tiny café workbook, laid out so that a
single screenshot can show three different kinds of reference at once:

    B2         a bare cell, with no name at all
    D2:F5      the ``Sales`` named range, a header-topped table
    H2:I8      the ``Hours`` named range, two columns of name and value

That mixture is the point of the prop. ``video/quarter.toml`` addresses B2 by coordinate
and the other two by name, and the animation annotates the sheet to show the difference —
so B2 must genuinely have no defined name, or the demonstration would be making a claim
the workbook contradicts.

The gaps matter too. Columns A, C and G are empty and narrow, which keeps the three
ranges visually separate on screen without any of them touching the sheet's edge, and
row 1 is left clear for the animation to hang its name badges in.

A generated prop rather than a committed mystery: the workbook is a binary, so without
this script a change to it — deleting a defined name, say — is invisible in a diff.

Run from the project root:  ``python scripts/create_quarter.py``
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName

VIDEO = Path(__file__).resolve().parent.parent / "video"
XLSX = VIDEO / "quarter.xlsx"

TITLE = "Riverside Cafe"
SALES = [
    ("Region", "Q1", "Q2"),
    ("North", 1200, 1350),
    ("South", 980, 1010),
    ("East", 1440, 1390),
]
HOURS = [
    ("Monday", "07:00-20:00"),
    ("Tuesday", "07:00-20:00"),
    ("Wednesday", "07:00-20:00"),
    ("Thursday", "07:00-20:00"),
    ("Friday", "07:00-22:00"),
    ("Saturday", "09:00-17:00"),
    ("Sunday", "Closed"),
]
# A, C and G are the spacer columns that hold the three ranges apart on screen.
WIDTHS = {"A": 3, "B": 22, "C": 3, "D": 12, "E": 10, "F": 10, "G": 3, "H": 14, "I": 16}
BOLD = Font(bold=True)


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Quarter"

    ws["B2"] = TITLE
    ws["B2"].font = BOLD

    for r, row in enumerate(SALES, start=2):
        for c, value in enumerate(row, start=4):
            cell = ws.cell(row=r, column=c, value=value)
            if r == 2:                      # the heading row `records` will consume
                cell.font = BOLD

    for r, (day, hours) in enumerate(HOURS, start=2):
        ws.cell(row=r, column=8, value=day)
        ws.cell(row=r, column=9, value=hours)

    for column, width in WIDTHS.items():
        ws.column_dimensions[column].width = width

    # Sales and Hours are named; B2 deliberately is not.  See the module docstring.
    for name, ref in [("Sales", "Quarter!$D$2:$F$5"), ("Hours", "Quarter!$H$2:$I$8")]:
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    VIDEO.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    print(f"Wrote workbook: {XLSX}")


if __name__ == "__main__":
    build()
