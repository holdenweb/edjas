"""Read an OpenDocument spreadsheet (``.ods``) using only the standard library.

A great deal of public-sector data is published as ODS rather than xlsx -- the MHCLG
local-authority revenue outturn tables, for instance, are ODS-only -- so a tool for
getting data out of government spreadsheets that cannot open one is missing its
audience.

Rather than teach every part of the expression language a second file format, an ODS
file is parsed here and presented as an ordinary :class:`openpyxl.Workbook`. Everything
downstream -- references, named ranges, pipelines, functions -- then works unchanged.
An ODS file is a zip of XML, so ``zipfile`` and ``xml.etree`` are all that is required
and EDJAS gains no new dependency.

Only values are read, never styles: that matches the ``data_only=True`` reading of xlsx
files, where a formula cell yields the value the application cached rather than the
formula text.
"""

import datetime as dt
import zipfile
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

__all__ = ["is_ods", "load_workbook"]

TABLE = "{urn:oasis:names:tc:opendocument:xmlns:table:1.0}"
OFFICE = "{urn:oasis:names:tc:opendocument:xmlns:office:1.0}"
TEXT = "{urn:oasis:names:tc:opendocument:xmlns:text:1.0}"

# Excel's grid, which is also openpyxl's limit. ODS pads every row out to the full width
# and the sheet out to the full height using repeat counts, so a file of a few hundred
# real rows routinely claims 1,048,576 of them. Those runs are empty and are skipped
# rather than materialised.
MAX_ROW = 1048576
MAX_COLUMN = 16384

# A run of *non-empty* cells does have to be materialised, and the grid above allows
# seventeen billion of them. A few hundred bytes of XML could therefore ask for more
# memory than the machine has, so a budget is enforced and exceeding it is an error
# rather than a hang.
MAX_CELLS = 5_000_000

# ODF wraps rows in these when a sheet has print titles or outline grouping. Their
# contents are ordinary rows of the sheet, and groups may nest.
ROW_GROUPS = (
    f"{TABLE}table-header-rows",
    f"{TABLE}table-rows",
    f"{TABLE}table-row-group",
)


def is_ods(path):
    """True if ``path`` names an OpenDocument spreadsheet, by extension."""
    return str(path).lower().endswith(".ods")


def _paragraph(node):
    """The text of one paragraph, including ODF's explicit whitespace elements.

    ODS does not store runs of spaces literally: ``<text:s text:c="4"/>`` stands for
    four of them, with tabs and line breaks similarly encoded. Plain ``itertext`` drops
    all three, which would strip the leading spaces that government workbooks use to
    indent subordinate rows -- turning an indented line item into something that looks
    like a heading.
    """
    parts = [node.text or ""]
    for child in node:
        if child.tag == f"{TEXT}s":
            parts.append(" " * int(child.get(f"{TEXT}c", 1)))
        elif child.tag == f"{TEXT}tab":
            parts.append("\t")
        elif child.tag == f"{TEXT}line-break":
            parts.append("\n")
        else:
            parts.append(_paragraph(child))  # a span, link or other inline wrapper
        parts.append(child.tail or "")
    return "".join(parts)


def _text(cell):
    """The visible text of a cell: its paragraphs, joined by newlines.

    Only ``text:p`` children count. A cell may also carry an ``office:annotation``
    (a comment, holding its author and date) or an anchored ``draw:frame``, and neither
    is part of the value -- openpyxl keeps a comment off ``cell.value`` too.
    """
    paragraphs = [_paragraph(p) for p in cell if p.tag == f"{TEXT}p"]
    if paragraphs:
        return "\n".join(paragraphs)
    return cell.get(f"{OFFICE}string-value")


def _cell_value(cell):
    """Convert one ``table:table-cell`` to a Python value, or None if it is empty.

    A formula cell carries both ``table:formula`` and the cached result; the result is
    what is read, so a total or an average extracts as the number it evaluated to.
    """
    kind = cell.get(f"{OFFICE}value-type")
    if kind is None:
        return None
    if kind in ("float", "percentage", "currency"):
        return _number(cell.get(f"{OFFICE}value"))
    if kind == "boolean":
        return cell.get(f"{OFFICE}boolean-value") == "true"
    if kind == "date":
        return _date(cell.get(f"{OFFICE}date-value"))
    if kind == "time":
        return _duration(cell.get(f"{OFFICE}time-value"))
    return _text(cell)


def _number(value):
    """A numeric cell's value; whole numbers read as ints, as they do from xlsx."""
    if value is None:
        return None
    try:
        number = float(value)
    except ValueError:
        return value  # keep the text rather than lose the cell to a malformed attribute
    return int(number) if number.is_integer() else number


def _date(value):
    if not value:
        return None
    try:
        return dt.datetime.fromisoformat(value)
    except ValueError:
        return value


def _duration(value):
    """Parse an ISO-8601 duration such as ``PT08H30M00S`` into a ``time``.

    A duration is a length of time, not a time of day, so anything reaching 24 hours
    has no ``time`` to be converted to. Rather than wrap it -- which would turn 26h30m
    into a plausible, wrong 02:30 -- the original text is returned unchanged.
    """
    if not value or not value.startswith("PT"):
        return value
    digits, parts = "", {}
    for character in value[2:]:
        if character.isdigit() or character == ".":
            digits += character
        else:
            parts[character] = float(digits or 0)
            digits = ""
    hours = parts.get("H", 0)
    if hours >= 24:
        return value
    seconds = parts.get("S", 0)
    return dt.time(
        int(hours),
        int(parts.get("M", 0)),
        int(seconds),
        int(round((seconds % 1) * 1_000_000)),
    )


def _row_values(row):
    """Return ``[(column, value), ...]`` for the non-empty cells of one row.

    ``table:number-columns-repeated`` compresses runs of identical cells, so the column
    a cell lands in depends on every repeat count before it. Getting this wrong shifts
    an entire sheet sideways, which is the classic way of misreading ODS.
    ``table:covered-table-cell`` marks a cell hidden beneath a merge: it holds no value
    of its own but still occupies its column.
    """
    values = []
    column = 0
    for cell in row:
        if not cell.tag.startswith(TABLE):
            continue
        if cell.tag not in (f"{TABLE}table-cell", f"{TABLE}covered-table-cell"):
            continue
        repeat = _count(cell.get(f"{TABLE}number-columns-repeated"))
        value = None
        if cell.tag == f"{TABLE}table-cell":
            value = _cell_value(cell)
        if value is not None:
            for offset in range(repeat):
                position = column + offset + 1
                if position > MAX_COLUMN:
                    break
                values.append((position, value))
        column += repeat
        if column > MAX_COLUMN:
            break
    return values


def _rows(element):
    """Yield a sheet's rows in document order, descending into ODF's row wrappers.

    Print titles and outline groups nest rows inside ``table:table-header-rows``,
    ``table:table-rows`` and ``table:table-row-group``, and groups may nest further.
    Taking only direct children would drop those rows *and* leave every row below them
    one place too high -- the vertical twin of the column shift ``_row_values`` guards
    against. Descending only into those three wrappers matters: a ``table:table`` may
    legally appear inside a cell, and a blanket search would hoist that sub-table's
    rows into this sheet.
    """
    for child in element:
        if child.tag == f"{TABLE}table-row":
            yield child
        elif child.tag in ROW_GROUPS:
            yield from _rows(child)


def _read_sheet(workbook, table, budget):
    """Add one ``table:table`` to ``workbook`` as a worksheet."""
    sheet = workbook.create_sheet(table.get(f"{TABLE}name"))
    row_number = 0
    for row in _rows(table):
        repeat = _count(row.get(f"{TABLE}number-rows-repeated"))
        values = _row_values(row)
        if not values:
            # An empty run, usually the padding to the bottom of the grid. Skip past it
            # without creating cells, so a sheet costs memory proportional to its data.
            row_number += repeat
            continue
        budget.spend(repeat * len(values))
        for _ in range(repeat):
            row_number += 1
            if row_number > MAX_ROW:
                return sheet
            for column, value in values:
                sheet.cell(row=row_number, column=column, value=value)
    return sheet


class _Budget:
    """A ceiling on the cells one file may materialise."""

    def __init__(self, path, limit=MAX_CELLS):
        self.path, self.left = path, limit

    def spend(self, cells):
        self.left -= cells
        if self.left < 0:
            raise ValueError(
                f"{self.path}: the spreadsheet asks for more than {MAX_CELLS:,} cells; "
                f"it is probably malformed"
            )


def _count(value):
    """A repeat count, defaulting to 1 and never negative or unparseable."""
    if value is None:
        return 1
    try:
        return max(0, int(value))
    except ValueError:
        return 1


def _sheet_of(end):
    """Split one end of an address into ``(sheet, ref)``, or ``(None, ref)``."""
    if "." not in end:
        return None, end
    sheet, _, ref = end.rpartition(".")
    # OpenFormula marks an absolute sheet with a leading '$' -- '$Sheet.$A$1' -- which
    # LibreOffice itself writes. Left in place it becomes part of the sheet's name, and
    # the reference can then never resolve.
    return sheet.lstrip("$").strip("'").replace("''", "'"), ref


def _split_address(address):
    """Split an ODS range address into ``(sheet, refs)``.

    ODS writes ``Sheet.$C$8:Sheet.$AH$430``, naming the sheet on both sides and using a
    dot; Excel and EDJAS write ``'Sheet'!$C$8:$AH$430``.

    A range spanning two sheets has no equivalent in EDJAS's model, which reads one
    rectangle from one sheet. Rather than quietly return its shape read off the first
    sheet, ``(None, ...)`` is returned so the caller can decline it -- the same choice
    made for multi-area references.
    """
    ends = address.split(":")
    sheet, start = _sheet_of(ends[0])
    if sheet is None:
        return None, address
    refs = [start]
    for end in ends[1:]:
        other, ref = _sheet_of(end)
        if other is not None and other != sheet:
            return None, address
        refs.append(ref)
    return sheet, ":".join(refs)


def _document_named_ranges(root):
    """The workbook-level named ranges, excluding any scoped to a single sheet.

    ODF allows a ``table:named-range`` inside a ``table:table``, where it names a range
    local to that sheet. Those share a namespace with nothing, so hoisting them to the
    workbook lets two sheets' identically-named ranges silently overwrite each other.
    """
    scoped = {
        id(named)
        for table in root.iter(f"{TABLE}table")
        for named in table.iter(f"{TABLE}named-range")
    }
    return [n for n in root.iter(f"{TABLE}named-range") if id(n) not in scoped]


def _add_defined_names(workbook, root):
    """Copy ODS named ranges across as openpyxl defined names."""
    for named in _document_named_ranges(root):
        name = named.get(f"{TABLE}name")
        address = named.get(f"{TABLE}cell-range-address")
        if not name or not address:
            continue
        sheet, refs = _split_address(address)
        if sheet is None:
            continue
        quoted = "'" + sheet.replace("'", "''") + "'"
        workbook.defined_names.add(DefinedName(name, attr_text=f"{quoted}!{refs}"))


def _add_tables(workbook, root):
    """Copy ODS database ranges across as Excel-style named tables, where possible.

    A database range is ODS's nearest equivalent of an Excel Table. Some carry generated
    names that Excel would reject, so a range that openpyxl declines is skipped rather
    than failing the whole read -- the data is still reachable by cell reference.
    """
    for database in root.iter(f"{TABLE}database-range"):
        name = database.get(f"{TABLE}name")
        address = database.get(f"{TABLE}target-range-address")
        if not name or not address:
            continue
        sheet_name, refs = _split_address(address)
        if sheet_name is None or ":" not in refs:
            continue
        try:
            sheet = workbook[sheet_name]
            sheet.add_table(Table(displayName=name, ref=refs.replace("$", "")))
        except (KeyError, ValueError):
            continue


def _content(path):
    """The ``content.xml`` of an ODS file, or a ValueError explaining why not.

    Everything that can be wrong with the container is turned into ValueError here.
    ``read_spec`` and the command line both handle that, so a damaged or encrypted file
    produces a sentence naming the file rather than a traceback.
    """
    try:
        with zipfile.ZipFile(path) as archive:
            return archive.read("content.xml")
    except KeyError as exc:
        raise ValueError(
            f"{path}: not an OpenDocument spreadsheet (it has no content.xml)"
        ) from exc
    except zipfile.BadZipFile as exc:
        raise ValueError(
            f"{path}: not a readable OpenDocument spreadsheet "
            f"(it is not a zip archive, or its contents are damaged)"
        ) from exc
    except RuntimeError as exc:  # zipfile raises this for an encrypted member
        raise ValueError(f"{path}: the spreadsheet could not be read ({exc})") from exc


def load_workbook(path):
    """Read ``path``, an OpenDocument spreadsheet, into an openpyxl workbook."""
    try:
        root = ElementTree.fromstring(_content(path))
    except ElementTree.ParseError as exc:
        raise ValueError(f"{path}: the spreadsheet's content is not valid XML ({exc})") from exc

    budget = _Budget(path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    # Only the sheets of the document itself: a table may also appear inside a cell.
    for table in _spreadsheet_tables(root):
        _read_sheet(workbook, table, budget)
    if not workbook.sheetnames:
        raise ValueError(f"{path}: the spreadsheet contains no sheets")
    _add_defined_names(workbook, root)
    _add_tables(workbook, root)
    return workbook


def _spreadsheet_tables(root):
    """The document's own sheets, in order, ignoring tables nested inside cells."""
    for body in root.iter(f"{OFFICE}spreadsheet"):
        for child in body:
            if child.tag == f"{TABLE}table":
                yield child
