"""The narrated introducer: a 30-second animation cut to a written voiceover.

An experiment alongside make_video.py rather than a replacement for it.  The two differ in
more than length.  This one opens on the workbook itself -- the wall-of-data shot that
motivated the first version is cut, to keep the whole thing inside thirty seconds -- and it
assembles the JSON rather than cross-fading into it: the cells stack, the labels align to
the top of each block, and only then does the punctuation arrive around them.

The script is in SCRIPT below, verbatim, with the second each line is spoken.  Everything
else follows from it: the timeline's anchors are those seconds and the subtitles are those
words, so a change to the script is a change to the film and the two cannot drift apart.

    uv run python video/narrated.py             # -> video/narrated.html
    uv run python video/narrated.py --timeline  # ...and print what the beats resolved to
"""

import json
import sys
import tomllib
from pathlib import Path

from openpyxl import load_workbook

from edjas import read_spec
from make_video import COLOURS, Timeline, reference, resolve

HERE = Path(__file__).resolve().parent

# The scripts, verbatim.  Each line gives the beat it cues, **how long that beat lasts**,
# the animation it calls for, and the words.
#
# Durations rather than clock times, because a duration is what you can actually measure:
# time a recorded line with a stopwatch, type the number in, and everything after it moves
# out of the way.  Clock times would mean renumbering the rest of the script to lengthen one
# sentence, which is the thing that stops anybody retiming anything.
SHORT = [
    ("sheet", 2.0, "spreadsheet fades in",
     "Here's an example spreadsheet,"),
    ("names", 2.0, "labels appear",
     "with a couple of named ranges."),
    ("spec", 6.0, "specifications fade in",
     "EDJAS lets you describe and name the data you want. For example"),
    ("title", 1.0, "title row highlights and begins snaking towards the spreadsheet",
     "a single cell"),
    ("sales", 1.0, "sales row highlights and begins snaking towards the spreadsheet",
     "a columnar table"),
    ("hours", 2.0, "hours row highlights and begins snaking towards the spreadsheet",
     "or a set of keys and values"),
    ("hold", 3.0, "all three animated threads reach their destination ranges, "
     "which highlight",
     "Your spreadsheet is never modified."),
    ("strip", 2.0, "background to specification and spreadsheet begins to fade, "
     "data and keys are surrounded by an outline",
     "The names and values"),
    ("stack", 3.0, "the data items are stacked vertically, and the labels are vertically "
     "aligned with the top of the relevant data item",
     "are packaged into a JSON data set."),
    ("flow", 4.0, "the outlines fade away, the text shrinks, and the remainder of the "
     "JSON appears",
     "Almost everything can consume JSON,"),
    ("box", 4.0, "the JSON document shrinks to a box labelled JSON, and the API, report, "
     "database, dashboard and web site icons appear over it and zoom out into position",
     "turning your spreadsheets into valuable data sources."),
]

# The long cut opens on the problem rather than on the example, spends real time on the
# specification language, and names each destination as its icon arrives.  Two liberties,
# both deliberate: "Almost everything can consume JSON, including APIs," is split at its
# comma, because the API icon arrives on the second half; and "including named ranges and
# off-sheet references" is given an animation of its own -- the badges the sentence is about
# swell while the threads are still crossing -- since the script leaves it to the words.
LONG = [
    ("nightmare", 4.0, "a crowded government workbook fills the frame",
     "Extracting data from spreadsheets can be a nightmare."),
    ("promise", 5.0, "the crowd holds, then cuts away",
     "Now there's a way to get just what you want from even the most complex designs."),
    ("sheet", 2.5, "spreadsheet fades in",
     "Here's an example spreadsheet,"),
    ("names", 2.5, "labels appear",
     "with a couple of named ranges."),
    ("spec", 7.0, "specifications fade in, a line at a time",
     "EDJAS uses a simple language to let you describe and name the data you want. "
     "That might be"),
    ("title", 1.5, "title row highlights and begins snaking towards the spreadsheet",
     "a single cell"),
    ("sales", 1.5, "sales row highlights and begins snaking towards the spreadsheet",
     "a columnar table"),
    ("hours", 2.0, "hours row highlights and begins snaking towards the spreadsheet",
     "or a set of keys and values"),
    ("conventions", 3.5, "the threads carry on crossing",
     "using exactly the same conventions you use in your spreadsheets,"),
    ("offsheet", 3.5, "the range badges and the cell address swell",
     "including named ranges and off-sheet references."),
    ("hold", 3.5, "all three animated threads reach their destination ranges, "
     "which highlight",
     "Your spreadsheet is never modified."),
    ("strip", 3.5, "background to specification and spreadsheet begins to fade, "
     "data and keys are surrounded by an outline",
     "Each name is then attached to the specified data."),
    ("stack", 4.0, "the data items are stacked vertically, and the labels are vertically "
     "aligned with the top of the relevant data item",
     "are packaged into a JSON data set."),
    ("flow", 4.5, "the outlines fade away, the text shrinks, and the remainder of the "
     "JSON appears",
     "Almost everything can consume JSON,"),
    ("apis", 2.0, "the JSON document shrinks to a box labelled JSON, and the API icon "
     "appears over it and zooms out into position",
     "including APIs,"),
    ("reports", 1.6, "the report icon appears over the JSON block and zooms out",
     "reports"),
    ("databases", 1.8, "the database icon appears over the JSON block and zooms out",
     "databases"),
    ("dashboards", 1.8, "the dashboard icon appears over the JSON block and zooms out",
     "dashboards and"),
    ("web", 2.0, "the web site icon appears over the JSON block and zooms out",
     "web pages"),
    ("sources", 5.0, "the destinations hold",
     "This turns your spreadsheets into data sources to drive your business processes."),
    ("install", 7.0, "an overlay fades in showing sample pip and uv installation commands",
     "Best of all, it's open source so you can build it into your business processes "
     "wherever it's needed!"),
]

# The knobs.  Every value here reaches the page as a CSS custom property, so a change is one
# line and takes effect everywhere the property is used -- and can be tried live in the
# browser console before being written down:
#
#     document.documentElement.style.setProperty('--icon-size', '180px'); seek(30)
#
# Anything that is a structural fact rather than a choice -- the 1280x720 frame, where the
# JSON lands, which cell a name covers -- is deliberately not here: those are measured or
# derived, and a knob for them would only be a way to make them disagree with the film.
LOOK = {
    # where the opening arrangement sits
    "sheet-left": "500px", "sheet-top": "150px",
    "spec-left": "32px", "spec-top": "236px", "spec-width": "400px",

    # Lines.  Thick enough to read in a contact-sheet tile, not only at full size: 2.5px is
    # perfectly legible in the frame and disappears entirely at a third of it, which is how
    # an arrow that was being drawn correctly came to be reported as missing.
    "thread-weight": "5px",
    "ray-weight": "6px", "ray-colour": "#8a7fa6",
    "feed-weight": "10px", "feed-colour": "#5b5170",
    "head-size": "3.2",                          # arrowhead, in stroke widths

    # the JSON, the box it becomes, and the stacked arrangement on the way
    "json-size": "17px", "json-ink": "#1e1a2b", "json-paper": "#e8e2f5",
    "box-size": "26px", "stack-size": "20px",

    # the destinations
    "icon-size": "138px", "icon-weight": "0.95", "icon-label": "18px",

    # the workbook the JSON is fed from: the real grid, scaled down
    "mini-top": "552px", "mini-scale": "0.34", "mini-w": "700px", "mini-h": "276px",

    # the words on screen
    "sub-size": "27px", "sub-bottom": "38px", "install-size": "30px",
}


def look_css(look):
    """The knobs as custom properties, for the top of the stylesheet."""
    return "  :root {\n" + "".join(f"    --{k}: {v};\n" for k, v in look.items()) + "  }"


INSTALL = ["pip install edjas", "uv add edjas"]
ICON_CUES = ["apis", "reports", "databases", "dashboards", "web"]


def cues(script):
    """Where each line starts, and the total runtime.

    A line normally begins where the one before it ended, so timing a recording means typing
    in how long each line took and letting the rest move out of the way.  A fifth element
    pins a line to an absolute second instead, for a beat that has to land on the clock
    whatever precedes it -- everything after a pin flows on from there, so the two notations
    mix freely.  Overrunning a pin is an error rather than a silent overlap.
    """
    at, out = 0.0, {}
    for entry in script:
        beat, secs = entry[0], entry[1]
        pinned = entry[4] if len(entry) > 4 else None
        if pinned is not None:
            if pinned < round(at, 3) - 1e-9:
                raise ValueError(f"{beat!r} is pinned to {pinned}s, but the lines before it "
                                 f"already run to {round(at, 3)}s")
            at = float(pinned)
        out[beat] = round(at, 3)
        at = round(at + secs, 3)
    return out, round(at, 3)


# API, report, database, dashboard, web site -- in the order the script names them, with the
# offset from the centre each travels to.  Stroke-drawn on currentColor so they inherit.
#
# At three times the size these need real room, and the lower middle of the frame is spoken
# for twice over -- by the subtitle, and by the spreadsheet the JSON is fed from.  So the
# destinations take the top corners, the sides and the top, and everything below the box
# reads downwards: box, feed arrow, spreadsheet, words.
ICONS = [
    ("An API", -440, -170,
     '<path d="M9 6 4 12l5 6"/><path d="M15 6l5 6-5 6"/>'),
    ("A report", 440, -170,
     '<rect x="5" y="3" width="14" height="18" rx="2"/><path d="M8.5 8h7M8.5 12h7M8.5 16h4"/>'),
    ("A database", -490, 55,
     '<ellipse cx="12" cy="6" rx="7" ry="3"/><path d="M5 6v12c0 1.7 3.1 3 7 3s7-1.3 7-3V6"/>'
     '<path d="M5 12c0 1.7 3.1 3 7 3s7-1.3 7-3"/>'),
    ("A dashboard", 490, 55,
     '<path d="M4 17a8 8 0 1 1 16 0"/><path d="M12 17l3.5-4.5"/><path d="M4 17h16"/>'),
    ("A web site", 0, -265,
     '<circle cx="12" cy="12" r="8"/><path d="M4 12h16"/>'
     '<path d="M12 4a12 12 0 0 1 0 16 12 12 0 0 1 0-16"/>'),
]


def timeline(script):
    """The beats, taken straight off the script's cues.

    An anchor is the second its line is spoken, so retiming the film means retiming the
    script -- which is the right way round when the words come first.  Scenes a script does
    not cue are switched off rather than removed: hasOpener and hasInstall reach seek() as
    0 and those layers stay at zero opacity throughout, so one animation renders either cut
    without either carrying the other's machinery.
    """
    CUE, runtime = cues(script)
    t = Timeline()
    t.end = round(runtime + 2.0, 3)

    # the crowded workbook the long cut opens on, and the cut away from it
    t.hasOpener = 1 if "nightmare" in CUE else 0
    t.denseAt = CUE.get("nightmare", 0.0)
    t.denseFor = 1.2
    # The two overlap: run them back to back and there is a blank frame between the crowd
    # leaving and the example arriving, which reads as a fault rather than as a cut.
    t.denseOutAt = CUE["sheet"] - 1.6 if t.hasOpener else 0.0
    t.denseOutFor = 1.4

    t.sheetAt = CUE["sheet"] - (1.2 if t.hasOpener else -0.2)
    t.sheetFor = 1.4

    t.namesAt = CUE["names"]                     # "with a couple of named ranges."
    t.namesFor = 1.2

    # The sentence that explains the whole idea runs six or seven seconds, so the panel
    # arrives and then names itself a line at a time rather than landing whole and leaving
    # the screen still while the words finish.  The stagger is whatever fits the gap.
    t.specAt = CUE["spec"]
    t.specFor = 1.0
    t.specRowAt = t.specAt + 0.6
    t.specRowPer = round((CUE["title"] - t.specRowAt - 0.3) / 3, 3)
    t.specRowFor = 0.9
    t.specRowsTo = t.specRowAt + 2 * t.specRowPer + t.specRowFor

    # "a single cell / a columnar table / or a set of keys and values", a row at a time
    t.threadAt = CUE["title"]
    t.threadPer = round(CUE["sales"] - CUE["title"], 3)
    t.threadTo = CUE["hold"]                     # and a shared arrival
    t.liftFor = 0.5                              # a specification row lighting up
    t.litFor = 0.6                               # a range highlighting as its thread lands

    # "including named ranges and off-sheet references" -- so they swell as it is said
    t.hasEmph = 1 if "offsheet" in CUE else 0
    t.emphAt = CUE.get("offsheet", 0.0)
    t.emphFor = 1.0

    t.stripAt = CUE["strip"]
    t.stripFor = 1.4
    t.showAt = t.stripAt - 0.02                  # the flyers take over just before
    t.outlineAt = t.stripAt + 0.3
    t.outlineFor = 0.9

    t.stackAt = CUE["stack"]                     # "are packaged into a JSON data set."
    t.stackFor = round(min(2.6, CUE["flow"] - CUE["stack"] - 0.4), 3)
    t.stackTo = t.stackAt + t.stackFor

    t.flowAt = CUE["flow"]                       # "Almost everything can consume JSON,"
    t.destAt = CUE["apis"] if "apis" in CUE else CUE["box"]   # when destinations start
    t.flowFor = round(min(2.8, t.destAt - t.flowAt - 0.6), 3)
    t.flowTo = t.flowAt + t.flowFor
    # Where the script names the destinations one at a time, the box has to have formed and
    # been fed before the first of them is spoken, so it goes as soon as the JSON is whole.
    t.boxAt = CUE["box"] if "box" in CUE else round(t.flowTo + 0.2, 3)
    t.outlineOutFor = 0.8
    t.synAt = t.flowAt + 0.9                     # the remainder of the JSON arrives
    t.synFor = 1.8
    t.cardAt = t.flowAt + 1.0                    # the card is solid before anything lands
    t.cardTo = t.cardAt + 1.2
    t.handFor = 0.7                              # flyers out as the real tokens come in

    # The destinations are simply there, at the size and place they finish, as the JSON
    # becomes a box -- nothing travels.  What arrives afterwards is the flow: the spreadsheet
    # the data came from, a thick arrow feeding the box, and then one arrow per destination
    # as the words reach it.
    t.boxFor = 1.6
    t.iconsAt = t.boxAt + 0.4
    t.iconsFor = 1.0
    t.miniAt = t.boxAt + 0.6                     # the spreadsheet it all came out of
    t.miniFor = 0.8
    t.feedAt = t.boxAt + 0.8
    t.feedFor = 0.7
    t.rayFor = 0.7                               # one destination arrow drawing
    t.rayAt = t.feedAt + t.feedFor               # where they start when not cued...
    t.rayPer = 0.5                               # ...and how far apart

    # an overlay of installation commands, on the long cut only
    t.hasInstall = 1 if "install" in CUE else 0
    t.installAt = CUE.get("install", 0.0)
    t.installFor = 1.4

    if CUE["hours"] - CUE["sales"] != t.threadPer:
        raise ValueError("the three rows are not evenly spaced in the script; the animation "
                         "staggers them by one interval and would drift from the words")
    if t.specRowsTo > t.threadAt:
        raise ValueError(f"the specification is still arriving at {t.specRowsTo}s when the "
                         f"first row is due to light at {t.threadAt}s")

    beats = ["sheetAt", "namesAt", "specAt", "threadAt", "stripAt", "stackAt",
             "flowAt", "boxAt"]
    for earlier, later in zip(beats, beats[1:]):
        if getattr(t, later) <= getattr(t, earlier):
            raise ValueError(f"{later} ({getattr(t, later)}s) does not follow "
                             f"{earlier} ({getattr(t, earlier)}s)")
    last = ray_starts(CUE, t)[-1] + t.rayFor
    if last > t.end:
        raise ValueError(f"the last icon settles at {last}s, past the end at {t.end}s")
    return t


def ray_starts(CUE, t):
    """When each destination's arrow draws: on its own word where the script names them one
    by one, otherwise on a fixed stagger once the feed arrow has arrived."""
    if all(cue in CUE for cue in ICON_CUES):
        return [CUE[cue] for cue in ICON_CUES]
    return [round(t.rayAt + i * t.rayPer, 3) for i in range(len(ICONS))]


def subtitles(script, t):
    """The voiceover, on screen, so the animation can be timed against the words.

    Each line runs to the start of the next, which is how the script is written; the last
    holds to the end.  They are the reason the film exists in this shape, so they are shown
    by default and switched off from the transport when it is time to record.
    """
    CUE, _ = cues(script)
    starts = [CUE[beat] for beat, _, _, _ in script]
    ends = starts[1:] + [t.end]
    return [[words, start, end]
            for (_, _, _, words), start, end in zip(script, starts, ends)]


def blocks(book, ranges, rows):
    """What each reference selects, as a grid of (address, value) -- the map everything else
    is built from.

    The three shapes come from the specification's own expression forms, so the animation
    cannot describe a shape the spec does not ask for: a bare reference is one cell, `[...]`
    with `records` is a heading row over records, and `{...}` is two columns of pairs.
    """
    ws = load_workbook(book, data_only=True).active
    out = {}
    for key, a1 in ranges.items():
        top_left, bottom_right = a1.split(":")     # single-letter columns; this sheet is A-I
        c1, r1 = ord(top_left[0]), int(top_left[1:])
        c2, r2 = ord(bottom_right[0]), int(bottom_right[1:])
        grid_ = [[(f"{chr(c)}{r}", ws.cell(row=r, column=c - ord("A") + 1).value)
                  for c in range(c1, c2 + 1)] for r in range(r1, r2 + 1)]
        expr = rows[key]
        kind = ("records" if "records" in expr else
                "object" if expr.strip().startswith("{") else "scalar")
        out[key] = {"kind": kind, "grid": grid_}
    return out


def _cell(value):
    return "" if value is None else str(value)


def sheet_html(path, max_row=9, max_col=9, ids=True):
    """The workbook's cells, each value wrapped in a span of its own.

    The morph measures text, not the boxes text happens to sit in: a cell is 26px tall
    whatever the 14px inside it is doing, and scaling by the box would land the words at the
    wrong size.  The span gives all three waypoints -- sheet, stack, JSON -- the same kind of
    thing to measure.
    """
    ws = load_workbook(path, data_only=True).active
    letters = [chr(ord("A") + i) for i in range(max_col)]
    rows = ["<tr><th></th>" + "".join(f"<th>{c}</th>" for c in letters) + "</tr>"]
    for r in range(1, max_row + 1):
        cells = [f"<th>{r}</th>"]
        for c in range(1, max_col + 1):
            addr = f"{letters[c - 1]}{r}"
            value = _cell(ws.cell(row=r, column=c).value)
            # the shrunken copy at the end takes no ids: two elements answering to cB2
            # would be invalid, and getElementById would quietly pick whichever came first
            cell_id = f' id="c{addr}"' if ids else ""
            span_id = f' id="cv-{addr}"' if ids else ""
            span = f'<span class="cv"{span_id}>{value}</span>' if value else ""
            cells.append(f"<td{cell_id}>{span}</td>")
        rows.append("<tr>" + "".join(cells) + "</tr>")
    return f'<table class="sheet">{"".join(rows)}</table>' 


def json_html(parts):
    """The JSON, token by token, with every value carrying the cell it came from.

    A `tok` is a value that exists in the workbook and so can fly out of it; a `syn` is
    everything else -- braces, quotes, commas, indentation, and the record keys that repeat
    on the second and third rows.  Those are the "remainder of the JSON" the script has
    appearing last, which is exactly what they are.
    """
    def syn(text):
        return f'<span class="syn">{text}</span>'

    def tok(addr, text):
        return f'<span class="tok" id="j-{addr}">{text}</span>'

    def key(name):
        return f'<span class="tok jk" id="jk-{name}">{name}</span>'

    out = [syn("{")]
    keys = list(parts)
    for n, name in enumerate(keys):
        tail = "," if n < len(keys) - 1 else ""
        block = parts[name]
        rows = block["grid"]
        if block["kind"] == "scalar":
            addr, value = rows[0][0]
            out.append(syn('  "') + key(name) + syn('": "') + tok(addr, _cell(value))
                       + syn(f'"{tail}'))
        elif block["kind"] == "records":
            heads = rows[0]
            out.append(syn('  "') + key(name) + syn('": ['))
            for i, row in enumerate(rows[1:]):
                bits = []
                for (haddr, head), (addr, value) in zip(heads, row):
                    label = tok(haddr, head) if i == 0 else f'<span class="syn">{head}</span>'
                    quoted = isinstance(value, str)
                    bits.append(syn('"') + label + syn('": ' + ('"' if quoted else ''))
                                + tok(addr, _cell(value)) + syn('"' if quoted else ''))
                comma = "," if i < len(rows) - 2 else ""
                out.append(syn("    { ") + syn(", ").join(bits) + syn(" }" + comma))
            out.append(syn(f"  ]{tail}"))
        else:
            out.append(syn('  "') + key(name) + syn('": {'))
            for i, ((kaddr, kval), (vaddr, vval)) in enumerate(rows):
                comma = "," if i < len(rows) - 1 else ""
                out.append(syn('    "') + tok(kaddr, _cell(kval)) + syn('": "')
                           + tok(vaddr, _cell(vval)) + syn(f'"{comma}'))
            out.append(syn(f"  }}{tail}"))
    out.append(syn("}"))
    return "".join(f'<div class="jl">{line}</div>' for line in out)


def stack_html(parts):
    """The halfway house: the blocks stacked vertically, each label level with its own top.

    This is measured, not drawn -- it is laid out off-screen so the flyers can read real
    positions out of it, exactly as they read their finish out of the JSON.  Writing the
    intermediate arrangement as ordinary markup is much less trouble than computing it, and
    it stays right when the data changes shape.
    """
    out = []
    for name, block in parts.items():
        rows = "".join(
            "<tr>" + "".join(
                f'<td><span id="s-{addr}">{_cell(value)}</span></td>' for addr, value in row
            ) + "</tr>" for row in block["grid"])
        out.append(f'<div class="sblock" style="--c:{COLOURS[name]}">'
                   f'<span class="slab" id="sk-{name}">{name}</span>'
                   f"<table>{rows}</table></div>")
    return f'<div id="stackInner">{"".join(out)}</div>'


def opener(t):
    """The wall of real data the long cut opens on, or nothing at all for the short one."""
    if not t.hasOpener:
        return "", []
    from importlib.resources import as_file

    from make_video import dense_grid, opener_workbook

    with as_file(opener_workbook()) as slgfs:
        return dense_grid(slgfs, "Scotland"), load_workbook(
            slgfs, read_only=True).sheetnames[:26]


def build(script=SHORT, out="narrated.html", show_timeline=False):
    book, spec = HERE / "quarter.xlsx", HERE / "quarter.toml"
    CUE, _ = cues(script)
    t = timeline(script)
    data = read_spec(str(book), str(spec))
    rows = tomllib.loads(spec.read_text())["extract"]
    wb = load_workbook(book)
    labels = {k: reference(v) for k, v in rows.items()}
    ranges = {k: resolve(labels[k]["text"], wb) for k in rows}
    parts = blocks(book, ranges, rows)
    dense, tabs = opener(t)

    payload = {
        "grid": sheet_html(book),
        "dense": dense,
        "denseTabs": tabs,
        "install": INSTALL,
        "specRows": "".join(
            f'<div class="srow" id="s{k}" style="--c:{COLOURS[k]}">'
            f'<span class="k">{k}</span> = <span class="v">"{v}"</span></div>'
            for k, v in rows.items()),
        "json": json_html(parts),
        "stack": stack_html(parts),
        "cells": {k: [a for row in v["grid"] for a, _ in row] for k, v in parts.items()},
        "ranges": ranges,
        "labels": labels,
        "colours": COLOURS,
        "icons": [[n, dx, dy, d] for n, dx, dy, d in ICONS],
        "rayAt": ray_starts(CUE, t),
        "mini": sheet_html(book, ids=False),
        "look": LOOK,
        "T": t.values(),
        "subtitles": subtitles(script, t),
    }
    html = (TEMPLATE.replace("__LOOK__", look_css(LOOK))
                    .replace("__PAYLOAD__", json.dumps(payload)))
    (HERE / out).write_text(html, encoding="utf-8")
    print(f"wrote {HERE / out}")
    if show_timeline:
        print(t.listing())
    return data


def listing(script):
    """The script as it will be spoken: where each line starts, how long it has, and the
    words -- the table to hold a stopwatch against while recording."""
    CUE, runtime = cues(script)
    out = []
    for entry in script:
        beat, secs, animation, words = entry[:4]
        at, pin = CUE[beat], "=" if len(entry) > 4 and entry[4] is not None else " "
        out.append(f" {pin}{int(at // 60)}:{at % 60:04.1f}  +{secs:<4.1f} {beat:<11} {words}")
        out.append(f"                          {animation}")
    out.append(f"\n  a leading = marks a line pinned to the clock rather than "
               f"following the one before it")
    out.append(f"  {len(script)} lines, {runtime}s spoken "
               f"({int(runtime // 60)}:{runtime % 60:04.1f})")
    return "\n".join(out)


TEMPLATE = r"""<!doctype html>
<meta charset="utf-8"><title>EDJAS narrated introducer</title>
<style>
__LOOK__
  :root { --ink:#1c1a22; --rule:#d8d2e4; --paper:#e8e2f5; }
  * { box-sizing:border-box; }
  html,body { height:100%; overflow:hidden; }
  body { margin:0; background:#3a3348; font-family:system-ui,-apple-system,"Segoe UI",sans-serif;
         display:flex; flex-direction:column; align-items:center; justify-content:center; gap:10px; }
  #sizer { position:relative; flex:0 0 auto; }
  #fit { position:absolute; left:0; top:0; width:1280px; height:720px; transform-origin:0 0; }
  #stage { position:relative; width:1280px; height:720px; background:#fff; overflow:hidden;
           color:var(--ink); box-shadow:0 12px 60px rgba(0,0,0,.5); }
  .layer { position:absolute; inset:0; }

  table.sheet { border-collapse:collapse; font-size:14px; position:absolute;
               left:var(--sheet-left); top:var(--sheet-top); }
  table.dense { position:absolute; left:0; top:0; font-size:10px; }
  table.dense td { min-width:62px; max-width:62px; height:21px; padding:1px 5px;
                   overflow:hidden; white-space:nowrap; text-overflow:ellipsis; }
  table.dense th { height:17px; font-size:9px; width:26px; }
  #tabs { position:absolute; left:20px; top:660px; display:flex; gap:2px; max-width:1240px;
          overflow:hidden; }
  #tabs span { flex:0 0 auto; font-size:11px; padding:3px 9px; background:#eee9f5;
               border:1px solid var(--rule); border-bottom:none; color:#655c78; }
  #tabs span.on { background:#fff; font-weight:700; color:var(--ink); }
  .sheet th { background:#eee9f5; color:#655c78; font-weight:600; font-size:12px;
              border:1px solid var(--rule); width:34px; height:26px; }
  .sheet td { border:1px solid var(--rule); padding:3px 7px; min-width:58px; height:26px;
              white-space:nowrap; }

  #labels .rbox { position:absolute; border-radius:3px; outline-offset:1px; }
  #labels .rbox.name { outline:2px dashed var(--c); }
  #labels .rbox.address { outline:2px solid var(--c); }
  #labels .chip { position:absolute; height:19px; display:flex; align-items:center;
                  transform-origin:0 50%;
                  padding:0 7px; border-radius:5px; font-size:12px; font-weight:700;
                  white-space:nowrap; }
  #labels .chip.name { background:var(--c); color:#fff; }
  #labels .chip.address { background:#fff; color:var(--c); border:1.5px solid var(--c);
                          font-family:ui-monospace,"SF Mono",Menlo,monospace; }

  /* The panel's vertical position is set from the measured layout, not here: each row has to
     sit below the bottom of every range to its left, or its thread reaches the far side of
     the sheet only by cutting straight through one of the others. */
  #spec { position:absolute; left:var(--spec-left); top:var(--spec-top);
          width:var(--spec-width); background:#faf8fd;
          border:1px solid var(--rule); border-radius:10px; padding:16px 18px;
          font-family:ui-monospace,"SF Mono",Menlo,monospace; font-size:15px; }
  #spec .hdr { color:#8a7fa6; margin-bottom:8px; }
  .srow { padding:8px; border-radius:6px; white-space:nowrap; }
  .srow .k { color:var(--c); font-weight:700; }
  .srow .v { color:#3d3550; }

  #threads { position:absolute; inset:0; width:1280px; height:720px; pointer-events:none; }
  #threads path { fill:none; stroke-width:var(--thread-weight); stroke-linecap:round;
                  stroke-dasharray:var(--len); stroke-dashoffset:var(--len); }

  /* the staging layout: measured, never seen */
  /* Laid out where the stack is meant to appear and simply not painted, so the middle of
     every flyer's journey is a real position on the stage rather than a computed guess. */
  #stack { position:absolute; left:50%; top:50%; transform:translate(-50%,-50%);
           opacity:0; pointer-events:none; font-size:var(--stack-size); }
  #stack .sblock { display:flex; align-items:flex-start; gap:22px; margin-bottom:20px; }
  #stack .slab { font-family:ui-monospace,Menlo,monospace; font-weight:700; color:var(--c);
                 min-width:78px; line-height:30px; }
  #stack table { border-collapse:collapse; }
  #stack td { padding:0 16px 0 0; height:30px; white-space:nowrap; vertical-align:middle; }

  #json { position:absolute; left:50%; top:50%; transform:translate(-50%,-50%);
          font-family:ui-monospace,Menlo,monospace; font-size:var(--json-size);
          line-height:1.55; background:var(--json-ink); color:var(--json-paper);
          padding:20px 28px; border-radius:12px;
          white-space:pre; opacity:0; box-shadow:0 20px 60px rgba(0,0,0,.4); }
  #json .syn, #json .tok { opacity:0; }
  #json .jk { font-weight:700; }

  #box { position:absolute; left:50%; top:50%; transform:translate(-50%,-50%);
         background:var(--json-ink); color:#fff; border-radius:12px; padding:16px 30px;
         opacity:0; font-family:ui-monospace,Menlo,monospace; font-size:var(--box-size);
         font-weight:700;
         box-shadow:0 16px 44px rgba(0,0,0,.34); }

  #morph { pointer-events:none; opacity:0; }
  .fly { position:absolute; transform-origin:0 0; white-space:nowrap; display:flex;
         align-items:center; }
  .fly .ring { position:absolute; inset:-4px -7px; border-radius:6px; opacity:0;
               outline:2px solid var(--c); }
  .fly.key { font-family:ui-monospace,"SF Mono",Menlo,monospace; font-weight:700;
             color:var(--c); }

  #icons .dest { position:absolute; left:50%; top:50%; display:flex; flex-direction:column;
                 align-items:center; gap:10px; opacity:0; color:var(--ink); }
  /* Drawn at the size they finish at and scaled down for the journey, so the stroke is
     resolved once: growing a 46px icon threefold would give it a nine-pixel line. */
  #icons svg { width:var(--icon-size); height:var(--icon-size); fill:none;
               stroke:currentColor; stroke-width:var(--icon-weight);
               stroke-linecap:round; stroke-linejoin:round; }
  #icons .name { font-size:var(--icon-label); font-weight:600; }
  /* An <svg> is a replaced element: inset:0 leaves it at its intrinsic 300x150 and clips
     everything drawn outside that, exactly as #threads above says.  It needs real size. */
  #rays { position:absolute; inset:0; width:1280px; height:720px; pointer-events:none; }
  /* The child combinator matters.  A <marker> is a descendant of #rays, so "#rays path"
     selected its arrowhead as well and painted it fill:none with a stroke -- CSS beats the
     fill presentation attribute, so the heads came out as hollow chevrons.  The lines are
     its direct children; the heads are not, and are filled solid below. */
  #rays > path { fill:none; stroke:var(--ray-colour); stroke-width:var(--ray-weight);
                 stroke-linecap:round; }
  #rays > path.feed { stroke:var(--feed-colour); stroke-width:var(--feed-weight); }
  #rays marker path { fill:currentColor; stroke:none; }
  #rays #rayHead { color:var(--ray-colour); }
  #rays #feedHead { color:var(--feed-colour); }
  /* the workbook it all came out of, the real grid at about a third of the size */
  #mini .card { position:absolute; left:50%; top:var(--mini-top);
                transform:translate(-50%,-50%); overflow:hidden; opacity:0; background:#fff;
                width:calc(var(--mini-w) * var(--mini-scale));
                height:calc(var(--mini-h) * var(--mini-scale));
                border:1px solid var(--rule); border-radius:6px;
                box-shadow:0 8px 24px rgba(0,0,0,.18); }
  #mini .inner { position:absolute; left:0; top:0;
                 width:var(--mini-w); height:var(--mini-h);
                 transform:scale(var(--mini-scale)); transform-origin:0 0; }
  #mini .inner table.sheet { left:0; top:0; }

  #install { position:absolute; inset:0; display:grid; place-items:center; opacity:0;
             background:rgba(255,255,255,.82); }
  #install .card { background:#1e1a2b; border-radius:14px; padding:26px 40px;
                   box-shadow:0 20px 60px rgba(0,0,0,.4); display:grid; gap:12px; }
  #install code { font-family:ui-monospace,Menlo,monospace; font-size:var(--install-size);
                  color:#fff; }
  #install code::before { content:"$ "; color:#8a7fa6; }
  #sub { position:absolute; left:60px; right:60px; bottom:var(--sub-bottom);
         text-align:center; font-size:var(--sub-size); font-weight:600; color:var(--ink);
         opacity:0; line-height:1.3; }
  #scrub { display:flex; gap:12px; align-items:center; color:#fff; flex:0 0 auto;
           font-family:system-ui; font-size:13px; }
  #scrub input[type=range] { flex:1; }
  #scrub label { display:flex; gap:5px; align-items:center; cursor:pointer; }
</style>

<div id="sizer"><div id="fit"><div id="stage">
  <div class="layer" id="denseLayer"></div>
  <div class="layer" id="sheetLayer"></div>
  <div class="layer" id="labels"></div>
  <div class="layer"><div id="spec"><div class="hdr">[extract]</div><div id="specRows"></div></div></div>
  <svg class="layer" id="threads"></svg>
  <div class="layer"><div id="json"></div></div>
  <div class="layer"><div id="box">JSON</div></div>
  <div class="layer" id="mini"><div class="card"><div class="inner"></div></div></div>
  <svg class="layer" id="rays"></svg>
  <div class="layer" id="morph"></div>
  <div class="layer" id="icons"></div>
  <div class="layer" id="install"><div class="card"></div></div>
  <div id="stack"></div>
  <div id="sub"></div>
</div></div></div>
<div id="scrub">
  <button id="play">play</button>
  <input id="t" type="range" min="0" max="32" step="0.02" value="0">
  <span id="tv">0.00s</span>
  <label><input id="subs" type="checkbox" checked> subtitles</label>
</div>

<script>
const D = __PAYLOAD__;
const T = D.T, SUBS = D.subtitles;
const stage = document.getElementById('stage');
const KEYS = Object.keys(D.ranges);
document.getElementById('sheetLayer').innerHTML = D.grid;
document.getElementById('specRows').innerHTML = D.specRows;
document.getElementById('json').innerHTML = D.json;
document.getElementById('stack').innerHTML = D.stack;
document.getElementById('denseLayer').innerHTML = D.dense ? D.dense +
  '<div id="tabs">' + D.denseTabs.map((n,i)=>
     `<span class="${i===0?'on':''}">${n}</span>`).join('') + '</div>' : '';
document.querySelector('#mini .inner').innerHTML = D.mini;
document.querySelector('#install .card').innerHTML =
  D.install.map(c=>`<code>${c}</code>`).join('');
document.getElementById('icons').innerHTML = D.icons.map(([n,,,d])=>
  `<div class="dest"><svg viewBox="0 0 24 24">${d}</svg><span class="name">${n}</span></div>`
).join('');
window.T = T;

const clamp=(x,a=0,b=1)=>Math.max(a,Math.min(b,x));
const ease =x=>x<.5?4*x*x*x:1-Math.pow(-2*x+2,3)/2;
const ph=(t,a,b)=>ease(clamp((t-a)/(b-a)));
const win=(t,a,b,f=.35)=>Math.min(ph(t,a,a+f), 1-ph(t,b-f,b));
const rgb=h=>[1,3,5].map(i=>parseInt(h.slice(i,i+2),16));
const mix=(a,b,k)=>{const A=rgb(a),B=rgb(b);
  return '#'+A.map((v,i)=>Math.round(v+(B[i]-v)*k).toString(16).padStart(2,'0')).join('');};

let SCALE=1;
function fit(force){
  const pad=20, bar=46;
  SCALE = force || Math.min((innerWidth-pad*2)/1280, (innerHeight-pad*2-bar)/720);
  document.getElementById('fit').style.transform=`scale(${SCALE})`;
  const sz=document.getElementById('sizer');
  sz.style.width=(1280*SCALE)+'px'; sz.style.height=(720*SCALE)+'px';
  document.getElementById('scrub').style.width=(1280*SCALE)+'px';
}
window.fit=fit; addEventListener('resize',()=>fit());
const box=el=>{const s=stage.getBoundingClientRect(), r=el.getBoundingClientRect();
  return {x:(r.left-s.left)/SCALE, y:(r.top-s.top)/SCALE, w:r.width/SCALE, h:r.height/SCALE,
          cx:(r.left-s.left+r.width/2)/SCALE, cy:(r.top-s.top+r.height/2)/SCALE};};

/* Drop the panel until every row clears the ranges to its left.  A row level with the sales
   block can only reach the hours columns by crossing it, which reads as a mistake however
   neatly it is drawn -- so the layout is corrected from the measurements rather than tuned
   by hand in the stylesheet. */
function placeSpec(){
  const spec=document.getElementById('spec');
  let shift=0;
  KEYS.forEach((k,i)=>{
    const row=box(document.getElementById('s'+k));
    KEYS.slice(0,i).forEach(j=>{
      const b=box(document.getElementById('c'+D.ranges[j].split(':')[1]));
      shift=Math.max(shift, b.y+b.h+14-row.cy);
    });
  });
  if(shift>0) spec.style.top=(parseFloat(getComputedStyle(spec).top)+shift)+'px';
}

let TAGS=[];
function buildLabels(){
  const L=document.getElementById('labels'); L.innerHTML=''; TAGS=[];
  for(const key of KEYS){
    const [tl,br]=D.ranges[key].split(':');
    const a=box(document.getElementById('c'+tl)), b=box(document.getElementById('c'+br));
    const w=b.x+b.w-a.x, h=b.y+b.h-a.y, k=D.labels[key].kind;
    const g=document.createElement('div');
    g.style.cssText=`--c:${D.colours[key]};position:absolute;inset:0`;
    g.innerHTML=
      `<div class="rbox ${k}" style="left:${a.x}px;top:${a.y}px;width:${w}px;height:${h}px"></div>`+
      `<div class="chip ${k}" style="left:${a.x}px;top:${a.y-22}px">${D.labels[key].text}</div>`;
    L.appendChild(g); TAGS.push(g);
  }
}

const svg=document.getElementById('threads');
function buildThreads(){
  svg.innerHTML='';
  KEYS.forEach(key=>{
    const row=box(document.getElementById('s'+key));
    const [tl,br]=D.ranges[key].split(':');
    const a=box(document.getElementById('c'+tl)), b=box(document.getElementById('c'+br));
    /* Aim below anything already passed.  Entering at the middle of the range is the
       natural choice, but the curve sags on its way across, and a thread that clips the
       corner of the block above reads as a mistake however neatly it is drawn.  Dropping
       the aim point keeps it under the obstacle and still lands squarely on the range. */
    const floor=Math.max(...KEYS.map(o=>{
      if(o===key) return 0;
      const q=box(document.getElementById('c'+D.ranges[o].split(':')[1]));
      const p0=box(document.getElementById('c'+D.ranges[o].split(':')[0]));
      return (p0.x < a.x) ? q.y+q.h+18 : 0;            /* only what lies to the left */
    }), 0);
    const x1=row.x+row.w, y1=row.cy, x2=a.x;
    const y2=Math.min(Math.max((a.y+b.y+b.h)/2, floor), b.y+b.h-10);
    const p=document.createElementNS('http://www.w3.org/2000/svg','path');
    p.setAttribute('d',`M${x1},${y1} C${x1+80},${y1} ${x2-80},${y2} ${x2},${y2}`);
    p.setAttribute('stroke',D.colours[key]); p.id='p'+key;
    svg.appendChild(p);
    p.style.setProperty('--len', p.getTotalLength());
  });
}

/* Three measured positions per flyer: where it is on the sheet, where it stands in the
   stacked arrangement, and where its own text sits in the finished JSON.  None is computed
   -- the staging layout is real markup off-screen, so the middle of the journey is as
   truthful as either end. */
let FLY=[];
function buildFlyers(){
  const morph=document.getElementById('morph');
  morph.innerHTML=''; FLY=[];
  const j=document.getElementById('json');
  const wasO=j.style.opacity; j.style.opacity=1;
  const toks=[...j.querySelectorAll('.tok')];
  toks.forEach(el=>el.style.opacity=1);

  const add=(src,mid,dst,text,cls,colour)=>{
    if(!src||!mid||!dst) return;
    const a=box(src), b=box(mid), c=box(dst);
    const em=el=>parseFloat(getComputedStyle(el).fontSize);
    const f=document.createElement('div');
    f.className='fly '+cls;
    f.style.cssText=`--c:${colour};left:${a.x}px;top:${a.y}px;font-size:${em(src)}px`;
    f.innerHTML=`<span class="ring"></span><span class="txt">${text}</span>`;
    morph.appendChild(f);
    FLY.push({el:f, ring:f.querySelector('.ring'), txt:f.querySelector('.txt'),
              x0:a.x, y0:a.y, x1:b.x, y1:b.y, x2:c.x, y2:c.y,
              s1:em(mid)/em(src), s2:em(dst)/em(src),
              from:cls==='key'?colour:'#1c1a22',
              to:cls==='key'?mix(colour,'#ffffff',0.3):'#e8e2f5'});
  };

  KEYS.forEach(key=>{
    add(document.querySelector('#s'+key+' .k'), document.getElementById('sk-'+key),
        document.getElementById('jk-'+key), key, 'key', D.colours[key]);
    D.cells[key].forEach(addr=>{
      const cv=document.getElementById('cv-'+addr);      /* the value, not the cell holding
                                                            it: a td is padded, and starting
                                                            from its corner leaves the text
                                                            visibly doubled at the handover */
      if(cv) add(cv, document.getElementById('s-'+addr),
                 document.getElementById('j-'+addr), cv.textContent, 'val', D.colours[key]);
    });
  });
  toks.forEach(el=>el.style.opacity='');
  j.style.opacity=wasO;
}

/* Arrows from the JSON box out to each destination.  Measured, like everything else: each
   destination is put into its final position for a moment so its icon can be read off the
   page, rather than the offsets being trusted to say where the drawing ends up -- the icon
   sits above its label, so the two are not the same point. */
let RAYS=[], FEED=null;
function buildArrows(){
  const NS='http://www.w3.org/2000/svg', HEAD=parseFloat(D.look['head-size']);
  const rays=document.getElementById('rays');
  /* markerUnits defaults to strokeWidth, so one head definition serves both weights */
  rays.innerHTML='<defs>'+
    `<marker id="rayHead" viewBox="0 0 10 10" refX="8" refY="5" markerWidth="${HEAD}"`+
    ` markerHeight="${HEAD}" orient="auto"><path d="M0,0 L10,5 L0,10 z"/></marker>`+
    `<marker id="feedHead" viewBox="0 0 10 10" refX="8" refY="5" markerWidth="${HEAD}"`+
    ` markerHeight="${HEAD}" orient="auto"><path d="M0,0 L10,5 L0,10 z"/></marker>`+
    '</defs>';
  const b=box(document.getElementById('box'));

  /* the thick one: out of the spreadsheet, into the box */
  const m=box(document.querySelector('#mini .card'));
  const fp=document.createElementNS(NS,'path');
  fp.setAttribute('class','feed'); fp.setAttribute('marker-end','url(#feedHead)');
  rays.appendChild(fp);
  FEED={el:fp, x0:m.cx, y0:m.y-10, x1:b.cx, y1:b.y+b.h+16};

  /* and one apiece out to the destinations, which are already standing where they finish */
  RAYS=[];
  [...document.querySelectorAll('#icons .dest')].forEach((d,i)=>{
    const [,dx,dy]=D.icons[i];
    const saved=d.style.transform;
    d.style.transform=`translate(-50%,-50%) translate(${dx}px,${dy}px)`;
    const g=box(d.querySelector('svg'));
    const dr=box(d);                                   /* the icon and its label together */
    d.style.transform=saved;
    const vx=g.cx-b.cx, vy=g.cy-b.cy, span=Math.hypot(vx,vy);
    const ux=vx/span, uy=vy/span;
    /* Leave the card at its edge and stop where the ray meets the destination -- the whole
       destination, not just the icon.  The one directly above the box has its label between
       the two, and an arrow aimed at the icon would run straight through the words. */
    const out=Math.min(Math.abs((b.w/2+16)/(ux||1e-6)), Math.abs((b.h/2+16)/(uy||1e-6)));
    const pad=14;
    const x0=dr.x-pad-b.cx, x1=dr.x+dr.w+pad-b.cx;
    const y0=dr.y-pad-b.cy, y1=dr.y+dr.h+pad-b.cy;
    const stop=Math.max(ux===0 ? -Infinity : (ux>0 ? x0/ux : x1/ux),
                        uy===0 ? -Infinity : (uy>0 ? y0/uy : y1/uy));
    const p=document.createElementNS(NS,'path');
    p.setAttribute('marker-end','url(#rayHead)');
    rays.appendChild(p);
    RAYS.push({el:p, x0:b.cx+ux*out, y0:b.cy+uy*out,
               x1:b.cx+ux*stop, y1:b.cy+uy*stop});
  });
}

/* An arrow grows towards where it is going rather than being revealed by a dash: a marker
   is drawn at the path's end whatever the dash pattern says, so the head would otherwise
   sit waiting at the far end from the first frame. */
function draw(a, k){
  a.el.setAttribute('d', `M${a.x0},${a.y0} `+
    `L${a.x0+(a.x1-a.x0)*k},${a.y0+(a.y1-a.y0)*k}`);
  a.el.style.opacity = k>0.04 ? Math.min(1,(k-0.04)*4) : 0;
}

function seek(t){
  document.getElementById('t').value=t;
  document.getElementById('tv').textContent=t.toFixed(2)+'s';

  /* the crowd the long cut opens on, then the workbook, then the names that belong to it */
  const strip=ph(t,T.stripAt,T.stripAt+T.stripFor);
  document.getElementById('denseLayer').style.opacity = T.hasOpener
    ? ph(t,T.denseAt,T.denseAt+T.denseFor)*(1-ph(t,T.denseOutAt,T.denseOutAt+T.denseOutFor))
    : 0;
  const sheetIn=ph(t,T.sheetAt,T.sheetAt+T.sheetFor)*(1-strip);
  document.getElementById('sheetLayer').style.opacity=sheetIn;
  const labels=document.getElementById('labels');
  labels.style.opacity=ph(t,T.namesAt,T.namesAt+T.namesFor)*(1-strip);
  /* "including named ranges and off-sheet references" -- so they swell as it is said.
     Only touched when a script asks for it: even scale(1) promotes the chip to a layer of
     its own and changes how its text is antialiased, which shows up as a differing frame. */
  if(T.hasEmph){
    const swell=1+0.28*ph(t,T.emphAt,T.emphAt+T.emphFor)*(1-strip);
    [...labels.querySelectorAll('.chip')].forEach(c=>c.style.transform=`scale(${swell})`);
  }

  const specIn=ph(t,T.specAt,T.specAt+T.specFor);
  const spec=document.getElementById('spec');
  spec.style.opacity=specIn*(1-strip);
  spec.style.transform=`translateX(${(specIn-1)*40}px)`;

  /* Each line of the specification arrives on its own, which gives the six seconds of
     "describe and name the data you want" something to do; then a row lights and its
     thread snakes across to the range it names. */
  KEYS.forEach((k,i)=>{
    const arrive=T.specRowAt+i*T.specRowPer;
    const written=ph(t,arrive,arrive+T.specRowFor);
    const at=T.threadAt+i*T.threadPer;
    const up=ph(t,at,at+T.liftFor)*(1-strip);
    const row=document.getElementById('s'+k);
    row.style.opacity=written;
    row.style.transform=`translateX(${(written-1)*16}px)`;
    row.style.background = up ? `rgba(255,255,255,${up})` : '';
    row.style.boxShadow  = up ? `0 ${8*up}px ${26*up}px rgba(0,0,0,${0.22*up})` : '';
    row.style.outline    = up ? `${2*up}px solid ${D.colours[k]}` : '';

    const p=document.getElementById('p'+k); if(!p) return;
    const drawn=ph(t, at, T.threadTo);
    p.style.strokeDashoffset=p.getTotalLength()*(1-drawn);
    p.style.opacity=1-ph(t,T.stripAt,T.stripAt+T.stripFor*0.7);

    const lit=ph(t,T.threadTo,T.threadTo+T.litFor)*(1-strip);
    D.cells[k].forEach(addr=>{
      const cell=document.getElementById('c'+addr); if(!cell) return;
      cell.style.boxShadow = lit ? `inset 0 0 0 ${3*lit}px ${D.colours[k]}` : '';
    });
  });

  /* the names and values leave the sheet, stack up, and become the JSON */
  const outline=ph(t,T.outlineAt,T.outlineAt+T.outlineFor)
                *(1-ph(t,T.flowAt,T.flowAt+T.outlineOutFor));
  const stack=ph(t,T.stackAt,T.stackTo);
  const flow=ph(t,T.flowAt,T.flowTo);
  const shrink=ph(t,T.boxAt,T.boxAt+T.boxFor);
  const hand=ph(t,T.flowTo-T.handFor,T.flowTo);      /* flyers out, real tokens in */
  document.getElementById('morph').style.opacity = t>=T.showAt ? 1-hand : 0;
  FLY.forEach(f=>{
    const x=f.x0+(f.x1-f.x0)*stack+(f.x2-f.x1)*flow;
    const y=f.y0+(f.y1-f.y0)*stack+(f.y2-f.y1)*flow;
    const s=1+(f.s1-1)*stack+(f.s2-f.s1)*flow;
    f.el.style.transform=`translate(${x-f.x0}px,${y-f.y0}px) scale(${s})`;
    f.ring.style.opacity=outline;
    f.txt.style.color=mix(f.from,f.to,flow);
  });

  const j=document.getElementById('json');
  j.style.opacity=ph(t,T.cardAt,T.cardTo)*(1-ph(t,T.boxAt+T.boxFor*0.2,T.boxAt+T.boxFor*0.75));
  j.style.transform=`translate(-50%,-50%) scale(${1-0.55*shrink})`;
  const syn=ph(t,T.synAt,T.synAt+T.synFor);
  [...j.querySelectorAll('.syn')].forEach(el=>el.style.opacity=syn);
  [...j.querySelectorAll('.tok')].forEach(el=>el.style.opacity=hand);

  /* ...which becomes a box, and feeds everything else */
  document.getElementById('box').style.opacity=ph(t,T.boxAt+T.boxFor*0.5,T.boxAt+T.boxFor);
  /* The destinations are simply present, at the size and place they finish; then the
     spreadsheet they all came from, the thick arrow feeding the box, and an arrow out to
     each destination as the words reach it. */
  const shown=ph(t,T.iconsAt,T.iconsAt+T.iconsFor);
  document.querySelector('#mini .card').style.opacity=ph(t,T.miniAt,T.miniAt+T.miniFor);
  if(FEED) draw(FEED, ph(t,T.feedAt,T.feedAt+T.feedFor));
  [...document.querySelectorAll('#icons .dest')].forEach((d,i)=>{
    const [,dx,dy]=D.icons[i];
    d.style.transform=`translate(-50%,-50%) translate(${dx}px,${dy}px)`;
    d.style.opacity=shown;
    if(RAYS[i]) draw(RAYS[i], ph(t,D.rayAt[i],D.rayAt[i]+T.rayFor));
  });

  document.getElementById('install').style.opacity =
    T.hasInstall ? ph(t,T.installAt,T.installAt+T.installFor) : 0;

  const line=SUBS.find(s=>t<s[2]) || SUBS[SUBS.length-1];
  const sub=document.getElementById('sub');
  sub.textContent=line[0];
  sub.style.opacity=document.getElementById('subs').checked ? win(t,line[1],line[2],.3) : 0;
}
window.seek=seek;

let playing=false, t0=0, base=0;
function frame(ts){ if(!playing) return;
  const t=base+(ts-t0)/1000; if(t>=T.end){ seek(T.end); playing=false; return; }
  seek(t); requestAnimationFrame(frame); }
document.getElementById('play').onclick=()=>{
  playing=!playing; base=+document.getElementById('t').value;
  if(playing) requestAnimationFrame(ts=>{t0=ts; frame(ts);}); };
document.getElementById('t').oninput=e=>{ playing=false; seek(+e.target.value); };
document.getElementById('subs').onchange=()=>seek(+document.getElementById('t').value);
document.getElementById('t').max=T.end;

addEventListener('load',()=>{
  fit(); placeSpec(); buildLabels(); buildThreads(); buildFlyers(); buildArrows(); seek(0);
});
</script>
"""

if __name__ == "__main__":
    long_cut = "--long" in sys.argv
    chosen = LONG if long_cut else SHORT
    if "--script" in sys.argv:
        print(listing(chosen))
    build(chosen,
          out="narrated-long.html" if long_cut else "narrated.html",
          show_timeline="--timeline" in sys.argv)
