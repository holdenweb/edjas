"""Build ``docs/site.xlsx``, the workbook the documentation site extracts itself from.

The site argues that a spreadsheet is a reasonable place to keep tabular content, and
that a small specification file can pull out whatever a given audience needs. This
workbook is that argument applied to the site: the gallery cards, the editorial columns
of the function table, and the two summary tables on the specification-language page are
all held here and extracted at build time by ``docs/site.toml``.

The columns that name things belonging to code -- function names, example names -- are
written from the live registries rather than restated, so the workbook cannot drift from
the software it describes. Tests assert the same thing from the other direction, in case
someone edits the committed file by hand.

Run from the project root:  ``python scripts/create_site_workbook.py``
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

import edjas_examples
from edjas.functions import DEFAULT_FUNCTIONS

DOCS = Path(__file__).resolve().parent.parent / "docs"
XLSX = DOCS / "site.xlsx"

# Editorial framing for each example: the shape it demonstrates and why a reader might
# care. Keyed by example name; the names themselves come from the live registry below.
GALLERY = {
    "retail": (
        "flat release",
        "A statistics release with a cover sheet, a contents sheet and a named Excel Table.",
        "Office for National Statistics",
    ),
    "slgfs": (
        "ledger",
        "A council revenue account: subservice lines rolling up to All Services.",
        "Scottish Government",
    ),
    "pesa": (
        "ledger",
        "A three-tier spending statement across six years of outturn and plans.",
        "HM Treasury",
    ),
    "dwp": (
        "ledger",
        "Estimate lines under Resource and Capital budgets, on a sheet named with a space.",
        "Department for Work and Pensions",
    ),
    "wales": (
        "ledger",
        "A revenue account whose columns mix pounds, percentages and per-head figures.",
        "Welsh Government",
    ),
    "cra": (
        "ledger",
        "Spending classified by international function, with dashes marking nil.",
        "HM Treasury",
    ),
    "nbs": (
        "large workbook",
        "Nineteen sheets, from which the specification takes one headline statement.",
        "Office for National Statistics",
    ),
}

# The prose columns of the function reference. The function *names* are taken from the
# registry, so a new built-in shows up here as a missing row rather than silently absent.
FUNCTIONS = {
    "records": ("[table]", "first row is headings; remaining rows become a list of objects"),
    "columns": ("[table]", "first row is headings; columns become a {heading: [values]} object"),
    "transpose": ("[table]", "swaps rows and columns"),
    "flatten": ("[table]", "flattens nested rows into a single list"),
    "keys": ("{object}", "the object's keys"),
    "values": ("{object}", "the object's values"),
    "items": ("{object}", "the object as [key, value] pairs"),
    "invert": ("{object}", "swaps keys and values"),
    "int": ("any", "coerces every value to a whole number"),
    "float": ("any", "coerces every value to a floating-point number"),
    "str": ("any", "coerces every value to text"),
    "round": ("any", "rounds floating-point values; takes the number of places (default 2)"),
    "isodate": ("any", "formats date and time values as ISO-8601 strings"),
}

FORMS = [
    ("ref", "the value of a single cell", 'title = "Summary!B2"'),
    ("[ref]", "the range as a list, or a list of row-lists", 'tags = "[Tags]"'),
    ("{ref}", "a two-column range as an object", 'hours = "{Hours}"'),
]

REF_KINDS = [
    ("named range", "Prices", "survives layout changes; the recommended form"),
    ("Excel Table name", "RevisionTriangles_Table1",
     "the table's whole range, header included, wherever it lives"),
    ("A1 range", "Summary!D3:E9", "explicit coordinates, optionally sheet-qualified"),
]


def _write(sheet, heading, rows):
    sheet.append(heading)
    for row in rows:
        sheet.append(list(row))


def main():
    workbook = Workbook()

    gallery = workbook.active
    gallery.title = "Gallery"
    _write(
        gallery,
        ["name", "shape", "teaser", "publisher"],
        # Iterating EXAMPLES, not GALLERY, is what keeps the two in step: a new example
        # with no editorial entry raises here rather than quietly vanishing from the site.
        [(name, *GALLERY[name]) for name in edjas_examples.EXAMPLES],
    )

    functions = workbook.create_sheet("Functions")
    _write(
        functions,
        ["name", "input", "result"],
        [(name, *FUNCTIONS[name]) for name in DEFAULT_FUNCTIONS],
    )

    forms = workbook.create_sheet("Forms")
    _write(forms, ["form", "produces", "example"], FORMS)

    kinds = workbook.create_sheet("RefKinds")
    _write(kinds, ["kind", "written as", "notes"], REF_KINDS)

    for sheet, last in (
        (gallery, "D"), (functions, "C"), (forms, "C"), (kinds, "C")
    ):
        name = f"{sheet.title}Table"
        ref = f"'{sheet.title}'!$A$1:${last}${sheet.max_row}"
        workbook.defined_names.add(DefinedName(name, attr_text=ref))

    DOCS.mkdir(parents=True, exist_ok=True)
    workbook.save(XLSX)
    print(f"wrote {XLSX}")


if __name__ == "__main__":
    main()
