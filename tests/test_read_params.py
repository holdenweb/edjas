import json
import re
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from edjas import json_default, read_spec
from edjas.read_params import parse_pipeline

DATA = Path(__file__).parent / "data"
FIXTURE_XLSX = DATA / "parameters.xlsx"
FIXTURE_TOML = DATA / "parameters.toml"


def make_workbook(tmp_path, cells, defined_names, extra_sheets=None):
    """Build a *data-only* workbook: {coord: value} on Sheet1, {name: ref} names.

    Cells hold data, never EDJAS markup. ``extra_sheets`` optionally adds further
    sheets as {sheet_name: {coord: value}} for cross-sheet references. Defined names
    are anchored to Sheet1.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for coord, value in cells.items():
        ws[coord] = value
    for sheet_name, sheet_cells in (extra_sheets or {}).items():
        sheet = wb.create_sheet(sheet_name)
        for coord, value in sheet_cells.items():
            sheet[coord] = value
    for name, ref in defined_names.items():
        wb.defined_names.add(DefinedName(name, attr_text=f"Sheet1!{ref}"))
    path = tmp_path / "data.xlsx"
    wb.save(path)
    return path


def make_spec(tmp_path, mapping, name="spec.toml"):
    """Write an [extract] TOML spec from {key: expression}.

    Expressions are emitted as TOML literal strings (single-quoted), so they may
    contain the double quotes EDJAS uses for string arguments.
    """
    lines = ["[extract]"]
    lines += [f"{key} = '{expr}'" for key, expr in mapping.items()]
    path = tmp_path / name
    path.write_text("\n".join(lines) + "\n")
    return path


# --- fixture (real data-only workbook + companion spec) ---------------------

@pytest.fixture(scope="module")
def params():
    return read_spec(FIXTURE_XLSX, FIXTURE_TOML)


def test_scalar(params):
    assert params["title"] == "EDJAS Demo"


def test_named_range_dict_hours(params):
    assert params["hours"]["Monday"] == "7:00 am - 8:00 pm"
    assert params["hours"]["Sunday"] == "Closed"
    assert len(params["hours"]) == 7


def test_named_range_dict_prices(params):
    assert params["prices"] == {"Tea": 3.25, "Coffee": 4.0, "Bacon Sandwich": 8.25}


def test_matrix_rows(params):
    assert params["hours_list"][0] == ["Monday", "7:00 am - 8:00 pm"]
    assert params["hours_list"][5] == ["Saturday", "9:00 am - 5:00 pm"]
    assert len(params["hours_list"]) == 7


def test_horizontal_vector(params):
    assert params["h_vector"] == ["the", "quick", "brown", "fox"]


def test_vertical_vector(params):
    assert params["v_vector"] == ["jumps", "over", "the", "lazy", "dog"]


def test_expected_keys_present(params):
    assert set(params) >= {"title", "hours", "prices", "hours_list", "h_vector", "v_vector"}


# --- pipeline parser -------------------------------------------------------

def test_parse_pipeline_bare_source():
    assert parse_pipeline("Sales") == ("Sales", [])


def test_parse_pipeline_chain_no_args():
    assert parse_pipeline("Grid | transpose | records") == (
        "Grid",
        [("transpose", []), ("records", [])],
    )


def test_parse_pipeline_quoted_arg_keeps_spaces_and_pipe():
    assert parse_pipeline('Words | join " | "') == (
        "Words",
        [("join", [(" | ", True)])],
    )


def test_parse_pipeline_mixed_args():
    assert parse_pipeline("Nums | clamp 0 10 Bounds") == (
        "Nums",
        [("clamp", [("0", False), ("10", False), ("Bounds", False)])],
    )


def test_parse_pipeline_rejects_multi_token_source():
    with pytest.raises(ValueError, match="single range reference"):
        parse_pipeline("a b | records")


def test_parse_pipeline_rejects_empty_stage():
    with pytest.raises(ValueError, match="Empty function stage"):
        parse_pipeline("Grid | transpose |")


def test_parse_pipeline_rejects_unterminated_string():
    with pytest.raises(ValueError, match="Unterminated string"):
        parse_pipeline('Words | join "oops')


# --- spec-driven extraction: the three forms --------------------------------

def test_scalar_extraction_active_sheet(tmp_path):
    xlsx = make_workbook(tmp_path, {"B2": "hello"}, {})
    spec = make_spec(tmp_path, {"greeting": "B2"})
    assert read_spec(xlsx, spec) == {"greeting": "hello"}


def test_scalar_extraction_named(tmp_path):
    xlsx = make_workbook(tmp_path, {"B2": "hi"}, {"Greeting": "$B$2"})
    spec = make_spec(tmp_path, {"g": "Greeting"})
    assert read_spec(xlsx, spec) == {"g": "hi"}


def test_object_extraction_with_pipeline(tmp_path):
    xlsx = make_workbook(
        tmp_path,
        {"D1": "Tea", "E1": "3", "D2": "Coffee", "E2": "4"},
        {"Prices": "$D$1:$E$2"},
    )
    spec = make_spec(tmp_path, {"prices": "{Prices | int}"})
    assert read_spec(xlsx, spec) == {"prices": {"Tea": 3, "Coffee": 4}}


def test_list_extraction_records(tmp_path):
    xlsx = make_workbook(
        tmp_path,
        {
            "D1": "Region", "E1": "Q1",
            "D2": "North", "E2": 100,
            "D3": "South", "E3": 200,
        },
        {"Sales": "$D$1:$E$3"},
    )
    spec = make_spec(tmp_path, {"sales": "[Sales | records]"})
    assert read_spec(xlsx, spec) == {
        "sales": [{"Region": "North", "Q1": 100}, {"Region": "South", "Q1": 200}]
    }


def test_chained_pipeline(tmp_path):
    xlsx = make_workbook(
        tmp_path,
        {
            "D1": "name", "E1": "alice", "F1": "bob",
            "D2": "age", "E2": 30, "F2": 25,
        },
        {"Grid": "$D$1:$F$2"},
    )
    spec = make_spec(tmp_path, {"people": "[Grid | transpose | records]"})
    assert read_spec(xlsx, spec) == {
        "people": [{"name": "alice", "age": 30}, {"name": "bob", "age": 25}]
    }


# --- spec-driven extraction: arguments --------------------------------------

def test_numeric_argument(tmp_path):
    xlsx = make_workbook(tmp_path, {"D1": 1, "E1": 2, "F1": 3}, {"Vec": "$D$1:$F$1"})
    spec = make_spec(tmp_path, {"vec": "[Vec | scale 10]"})
    out = read_spec(xlsx, spec, functions={"scale": lambda v, n: [x * n for x in v]})
    assert out == {"vec": [10, 20, 30]}


def test_quoted_argument(tmp_path):
    xlsx = make_workbook(
        tmp_path, {"D1": "the", "E1": "quick", "F1": "brown"}, {"Words": "$D$1:$F$1"}
    )
    spec = make_spec(tmp_path, {"csv": '[Words | join ", "]'})
    out = read_spec(xlsx, spec, functions={"join": lambda v, sep: sep.join(v)})
    assert out == {"csv": "the, quick, brown"}


def test_reference_argument(tmp_path):
    xlsx = make_workbook(
        tmp_path,
        {"D1": 1, "E1": 2, "F1": 3, "D2": 10, "E2": 20, "F2": 30},
        {"Vec": "$D$1:$F$1", "Other": "$D$2:$F$2"},
    )
    spec = make_spec(tmp_path, {"totals": "[Vec | add Other]"})
    out = read_spec(
        xlsx, spec, functions={"add": lambda v, other: [a + b for a, b in zip(v, other)]}
    )
    assert out == {"totals": [11, 22, 33]}


# --- spec-driven extraction: reference styles -------------------------------

def test_raw_range_source(tmp_path):
    xlsx = make_workbook(
        tmp_path,
        {"D1": 1, "E1": 2, "F1": 3, "D2": 4, "E2": 5, "F2": 6},
        {},
    )
    spec = make_spec(tmp_path, {"grid": "[D1:F2 | transpose]"})
    assert read_spec(xlsx, spec) == {"grid": [[1, 4], [2, 5], [3, 6]]}


def test_cross_sheet_raw_range_source(tmp_path):
    xlsx = make_workbook(
        tmp_path,
        {},
        {},
        extra_sheets={"Data": {
            "D1": 1, "E1": 2, "F1": 3,
            "D2": 4, "E2": 5, "F2": 6,
        }},
    )
    spec = make_spec(tmp_path, {"grid": "[Data!D1:F2 | transpose]"})
    assert read_spec(xlsx, spec) == {"grid": [[1, 4], [2, 5], [3, 6]]}


def test_cross_sheet_reference_argument(tmp_path):
    xlsx = make_workbook(
        tmp_path,
        {"D1": 1, "E1": 2, "F1": 3},
        {},
        extra_sheets={"Data": {"D1": 10, "E1": 20, "F1": 30}},
    )
    spec = make_spec(tmp_path, {"totals": "[D1:F1 | add Data!D1:F1]"})
    out = read_spec(
        xlsx, spec, functions={"add": lambda v, other: [a + b for a, b in zip(v, other)]}
    )
    assert out == {"totals": [11, 22, 33]}


# --- Excel reference edge cases --------------------------------------------

def _workbook_with_named_sheet(tmp_path, sheet_name, cells, defined_names):
    """Build a workbook whose data sheet has an arbitrary (e.g. spaced) name."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for coord, value in cells.items():
        ws[coord] = value
    for name, ref in defined_names.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))
    path = tmp_path / "named.xlsx"
    wb.save(path)
    return path


def test_named_range_on_sheet_with_spaces(tmp_path):
    """Excel quotes a spaced sheet name ('Sales Data'!$B$2); the quotes must be stripped."""
    xlsx = _workbook_with_named_sheet(
        tmp_path, "Sales Data", {"B2": "hi"}, {"Greeting": "'Sales Data'!$B$2"}
    )
    spec = make_spec(tmp_path, {"g": "Greeting"})
    assert read_spec(xlsx, spec) == {"g": "hi"}


def test_named_range_to_spaced_sheet_list(tmp_path):
    """A [range] via a named range also reaches a spaced sheet (quotes stripped)."""
    xlsx = _workbook_with_named_sheet(
        tmp_path, "Q1 Data", {"A1": 1, "B1": 2}, {"Row": "'Q1 Data'!$A$1:$B$1"}
    )
    spec = make_spec(tmp_path, {"row": "[Row]"})
    assert read_spec(xlsx, spec) == {"row": [1, 2]}


def test_sheet_name_with_embedded_apostrophe(tmp_path):
    """Excel doubles an embedded apostrophe inside the quotes ('Bob''s')."""
    xlsx = _workbook_with_named_sheet(
        tmp_path, "Bob's Data", {"C3": 42}, {"Answer": "'Bob''s Data'!$C$3"}
    )
    spec = make_spec(tmp_path, {"a": "Answer"})
    assert read_spec(xlsx, spec) == {"a": 42}


def test_multi_area_union_range_reports_clearly(tmp_path):
    """A union reference is unsupported and must fail with a clear message."""
    xlsx = _workbook_with_named_sheet(
        tmp_path, "Sheet", {"A1": 1, "B2": 2}, {"U": "Sheet!$A$1,Sheet!$B$2"}
    )
    spec = make_spec(tmp_path, {"u": "U"})
    with pytest.raises(ValueError, match="multi-area"):
        read_spec(xlsx, spec)


def test_defined_name_case_insensitive(tmp_path):
    """Excel matches defined names case-insensitively; so should EDJAS."""
    xlsx = make_workbook(tmp_path, {"B2": "hi"}, {"Title": "$B$2"})
    spec = make_spec(tmp_path, {"t": "title"})  # workbook defines "Title"
    assert read_spec(xlsx, spec) == {"t": "hi"}


def test_sheet_name_case_insensitive(tmp_path):
    """A sheet-qualified ref matches the sheet name case-insensitively too."""
    xlsx = make_workbook(tmp_path, {"D1": 1, "E1": 2}, {})
    spec = make_spec(tmp_path, {"row": "[sheet1!D1:E1]"})  # sheet is "Sheet1"
    assert read_spec(xlsx, spec) == {"row": [1, 2]}


def test_exact_case_defined_name_still_wins(tmp_path):
    """An exact-case match is preferred over the case-insensitive fallback."""
    xlsx = make_workbook(
        tmp_path, {"B2": "upper", "C3": "lower"},
        {"Item": "$B$2", "item": "$C$3"},
    )
    spec = make_spec(tmp_path, {"a": "Item", "b": "item"})
    assert read_spec(xlsx, spec) == {"a": "upper", "b": "lower"}


# --- formula cells (read with data_only=True) ------------------------------

def _formula_workbook(tmp_path, formula, cached=None):
    """Build a workbook with one formula cell (named ``Total``).

    openpyxl cannot write a cached formula result, so when ``cached`` is given we
    inject the ``<v>`` that Excel would store on recalculation, letting
    ``data_only=True`` read a genuine computed value.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"], ws["A2"] = 10, 20
    ws["A3"] = formula
    wb.defined_names.add(DefinedName("Total", attr_text="S!$A$3"))
    raw = tmp_path / "raw.xlsx"
    wb.save(raw)
    if cached is None:
        return raw
    with zipfile.ZipFile(raw) as zin:
        blobs = {n: zin.read(n) for n in zin.namelist()}
    sheet = next(n for n in blobs if re.search(r"xl/worksheets/sheet1\.xml$", n))
    f = formula.lstrip("=")
    xml = blobs[sheet].decode("utf-8").replace(f"<f>{f}</f>", f"<f>{f}</f><v>{cached}</v>")
    blobs[sheet] = xml.encode("utf-8")
    out = tmp_path / "cached.xlsx"
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in blobs.items():
            zout.writestr(name, data)
    return out


def test_formula_cell_yields_cached_value(tmp_path):
    """A formula cell reads as its cached computed value, not the formula text."""
    xlsx = _formula_workbook(tmp_path, "=A1+A2", cached=30)
    spec = make_spec(tmp_path, {"total": "Total"})
    assert read_spec(xlsx, spec) == {"total": 30}


def test_uncached_formula_reads_as_none(tmp_path):
    """A workbook never recalculated in Excel has no cache, so a formula is None."""
    xlsx = _formula_workbook(tmp_path, "=A1+A2")  # no cached value injected
    spec = make_spec(tmp_path, {"total": "Total"})
    assert read_spec(xlsx, spec) == {"total": None}


# --- errors and serialization ----------------------------------------------

def test_unknown_function_raises(tmp_path):
    xlsx = make_workbook(tmp_path, {"D1": 1, "E1": 2}, {"Vec": "$D$1:$E$1"})
    spec = make_spec(tmp_path, {"v": "[Vec | bogus]"})
    with pytest.raises(ValueError, match="Unknown EDJAS function 'bogus'"):
        read_spec(xlsx, spec)


def test_error_names_the_failing_key(tmp_path):
    """A failing expression is reported with its spec key and the expression."""
    xlsx = make_workbook(tmp_path, {"A1": 1}, {})
    spec = make_spec(tmp_path, {"good": "A1", "oops": "[Nope | records]"})
    with pytest.raises(ValueError, match=r"oops") as excinfo:
        read_spec(xlsx, spec)
    msg = str(excinfo.value)
    assert "oops" in msg and "Nope" in msg           # key and offending expression
    assert excinfo.value.__cause__ is not None       # original error preserved


def test_spec_without_extract_table_raises(tmp_path):
    xlsx = make_workbook(tmp_path, {"A1": "x"}, {})
    bad = tmp_path / "bad.toml"
    bad.write_text("[other]\nfoo = 'bar'\n")
    with pytest.raises(ValueError, match=r"must contain an \[extract\] table"):
        read_spec(xlsx, bad)


def test_scalar_pipeline_isodate(tmp_path):
    xlsx = make_workbook(tmp_path, {"B2": datetime(2026, 6, 29, 9, 30)}, {})
    spec = make_spec(tmp_path, {"opened": "B2 | isodate"})
    assert read_spec(xlsx, spec) == {"opened": "2026-06-29T09:30:00"}


def test_date_cell_serialises(tmp_path):
    xlsx = make_workbook(tmp_path, {"B2": datetime(2026, 6, 29, 9, 30)}, {})
    spec = make_spec(tmp_path, {"opened": "B2"})
    data = read_spec(xlsx, spec)
    assert data["opened"] == datetime(2026, 6, 29, 9, 30)
    assert json.loads(json.dumps(data, default=json_default)) == {
        "opened": "2026-06-29T09:30:00"
    }
