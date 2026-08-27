"""Generate the EDJAS introducer as a self-contained, seekable HTML animation.

Everything on screen is real: the grid is read from the workbook with openpyxl, the
specification is the actual .toml, and the JSON is what read_spec() returns. Nothing is
mocked up, so re-running this after a change to EDJAS produces a truthful video.

The animation is a pure function of time -- seek(t) computes every element's state from
t alone -- so it can be scrubbed by hand, played in a browser, or stepped frame by frame
by a headless browser for deterministic encoding.

    python video/make_video.py      # -> video/introducer.html

then open video/introducer.html in any browser, or drive it with video/contact.py and
video/check.py.  The page is self-contained and needs no server.
"""

import json
import re
import sys
import tomllib
from importlib.resources import as_file
from pathlib import Path

from openpyxl import load_workbook

from edjas import read_spec

try:  # the opener is a real published workbook, so it comes from the examples package
    from edjas_examples import DATA
except ModuleNotFoundError:  # pragma: no cover - a checkout-only tool, not shipped code
    raise SystemExit(
        "video/make_video.py needs the edjas-examples package: run `uv sync` "
        "in the repository root, then `uv run python video/make_video.py`."
    )

HERE = Path(__file__).resolve().parent
COLOURS = {  # Okabe-Ito: distinguishable under the common colour-vision deficiencies
    "title": "#D55E00",
    "sales": "#0072B2",
    "hours": "#009E73",
}
A1 = re.compile(r"^\$?[A-Z]{1,3}\$?[0-9]+(?::\$?[A-Z]{1,3}\$?[0-9]+)?$")


def reference(expr):
    """The reference a spec expression asks for, and whether it is an address or a name.

    This is what gets drawn on the sheet, so it has to be what the specification actually
    says -- strip the list/object brackets and any pipeline, and keep the rest verbatim.
    """
    ref = expr.strip().lstrip("[{").rstrip("]}").split("|")[0].strip()
    return {"text": ref, "kind": "address" if A1.match(ref) else "name"}


def resolve(ref, workbook):
    """Where that reference points, as a plain A1 range: a name is looked up, an address is not."""
    if ref in workbook.defined_names:
        ref = workbook.defined_names[ref].attr_text.split("!", 1)[1]
    a1 = ref.replace("$", "")
    return a1 if ":" in a1 else f"{a1}:{a1}"


class Timeline:
    """Named instants and durations, each fixed absolutely or against one already fixed.

    Reading a name that has not been fixed yet is an error, and so is fixing one twice.
    Between them those two rules mean the timeline reads top to bottom: whatever a line
    depends on is above it, nothing can refer forwards, and no cycle can be written down.
    Resolution is therefore just running the function.

    Values are rounded, because a chain of additions otherwise arrives at 26.159999999999997
    and every listing of the timeline is then unreadable.
    """

    def __init__(self):
        object.__setattr__(self, "_at", {})

    def __getattr__(self, name):
        try:
            return self._at[name]
        except KeyError:
            raise AttributeError(
                f"the timeline uses {name!r} before fixing it -- a point must be defined "
                f"above every line that refers to it"
            ) from None

    def __setattr__(self, name, value):
        if name in self._at:
            raise ValueError(f"the timeline fixes {name!r} twice "
                             f"(already {self._at[name]}s)")
        self._at[name] = round(float(value), 4)

    def values(self):
        return dict(self._at)

    def listing(self):
        return "\n".join(f"  {v:>7.2f}  {k}" for k, v in self._at.items())


def timeline():
    """Every instant and duration in the animation, in the order they happen.

    Anchors are absolute; everything else hangs off the beat before it.  That is what makes
    the video retimable: moving one anchor moves the whole of what follows, and the
    relationships that have to hold -- the flight starting after the growing finishes, the
    JSON shrinking only once it has finished assembling -- hold by construction rather than
    by being remembered.  Durations sit beside the instants they measure.
    """
    t = Timeline()
    t.end = 35                                   # how far the scrubber runs

    # the crowded opener calms down, and we cut to the clean workbook
    t.cutAt = 3.0
    t.cutFor = 1.4
    t.cutTo = t.cutAt + t.cutFor

    # the named ranges are the workbook's own property, so they arrive with the workbook
    t.namesAt = t.cutTo + 0.2
    t.namesFor = 1.0

    # the specification slides in from the left
    t.specAt = t.namesAt + 0.4
    t.specFor = 2.2
    t.specTo = t.specAt + t.specFor

    # the rows lift one by one, each bringing its own reference onto the sheet with it
    t.liftAt = t.specTo + 1.2
    t.liftPer = 0.35                             # the stagger between the three
    t.liftFor = 0.4
    t.chipFor = 0.8
    t.liftTo = t.liftAt + 2 * t.liftPer + t.liftFor

    # threads draw from staggered starts to a shared arrival, and the cells light
    t.threadAt = t.liftTo + 2.5
    t.threadPer = 0.4
    t.threadFor = 5.5
    t.threadTo = t.threadAt + t.threadFor
    t.threadLit = 0.85                           # a fraction of the draw, not an instant
    t.litOutFor = 0.35

    # the morph: the keys and values grow while everything else is stripped away
    t.morphAt = t.threadTo + 3.5
    t.showAt = t.morphAt - 0.02                  # flyers take over before anything moves
    t.growFor = 1.6
    t.padFor = 0.9
    t.chromeGone = t.morphAt + 1.4               # threads and lifted rows have gone
    t.stripAt = t.morphAt + 0.4                  # the sheet under the flyers starts to go
    t.stripFor = 1.6
    t.stripTo = t.stripAt + t.stripFor
    t.litTo = t.stripTo                          # lit cells go out as the sheet finishes

    # ...and then fly, arriving as the JSON
    t.flightAt = t.morphAt + t.growFor + 0.6     # only once the growing has finished
    t.flightFor = 2.2
    t.flightTo = t.flightAt + t.flightFor
    t.flightPer = 0.18
    t.padOutFor = 1.0
    t.padTo = t.flightAt + t.padOutFor
    t.cardAt = t.flightAt + 0.1
    t.cardFor = 0.9
    t.cardTo = t.cardAt + t.cardFor              # solid before the first flyer lands
    t.handLead = 0.1
    t.slotLead = 0.05
    t.handKey = 0.55                             # a key hands over more slowly than...
    t.handBlock = 0.40                           # ...a block of values, which is messier
    t.landsAt = t.flightTo + 2 * t.flightPer + t.handBlock   # the morph is over here

    # where the JSON goes next
    t.shrinkAt = t.landsAt + 0.44                # never before the JSON has assembled
    t.shrinkFor = 2.4
    t.shrinkTo = t.shrinkAt + t.shrinkFor
    t.fanAt = t.shrinkAt + 0.4
    t.fanFor = 1.0
    t.destAt = t.fanAt + 0.2
    t.destFor = 2.0
    t.destPer = 0.22

    # the close
    t.ctaAt = 32.9
    t.ctaFor = 1.2
    t.ctaTo = t.ctaAt + t.ctaFor

    # Everything above holds by construction except the order of the beats themselves: a
    # negative offset, or an anchor moved too far, can put one before the one it follows.
    beats = ["cutAt", "namesAt", "specAt", "liftAt", "threadAt", "morphAt",
             "flightAt", "shrinkAt", "fanAt", "ctaAt"]
    for earlier, later in zip(beats, beats[1:]):
        if getattr(t, later) <= getattr(t, earlier):
            raise ValueError(f"{later} ({getattr(t, later)}s) does not follow "
                             f"{earlier} ({getattr(t, earlier)}s)")
    if t.ctaTo > t.end:
        raise ValueError(f"the closing card ends at {t.ctaTo}s, past the end at {t.end}s")
    return t


def captions(t):
    """What is said on screen, anchored to the beats it belongs to.

    Each caption owns a window and a fade; win() keeps it at zero opacity outside that, so
    the one displayed is simply the first whose window has not closed.  Anchoring them here
    is what stops a retimed beat leaving its own caption behind.
    """
    return [
        ["Have you ever had to extract data from a spreadsheet like this?",
         0.3, t.cutAt + 0.4, 0.5],
        ["A cell.  A table.  A pair of columns.",
         t.threadAt + 0.5, t.morphAt - 0.4, 0.6],
        ["Your spreadsheet is never modified.",
         t.morphAt + 0.8, t.shrinkAt - 0.2, 0.6],
        ["JSON goes anywhere.",
         t.shrinkAt + 1.6, t.ctaAt - 0.3, 0.6],
    ]


def grid(path, max_row=9, max_col=9):
    """The workbook's real cells, as an HTML table with spreadsheet chrome."""
    ws = load_workbook(path, data_only=True).active
    letters = [chr(ord("A") + i) for i in range(max_col)]
    head = "".join(f"<th>{c}</th>" for c in letters)
    rows = [f"<tr><th></th>{head}</tr>"]
    for r in range(1, max_row + 1):
        cells = [f"<th>{r}</th>"]
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            cells.append(f'<td id="c{chr(ord("A")+c-1)}{r}">{"" if v is None else v}</td>')
        rows.append("<tr>" + "".join(cells) + "</tr>")
    return f'<table class="sheet">{"".join(rows)}</table>'


def json_html(data, records=2, pairs=3):
    """The real result, marked up so the animation can fly the keys and values into it.

    Every key carries an anchor (jk-*) and every value block another (jvb-*). seek() measures
    those boxes from the live layout and lands the specification's keys and the spreadsheet's
    values exactly on them, so the morph follows the real typography rather than hand-tuned
    coordinates -- change the font and the animation still converges.

    The output is abridged: 29 lines is taller than the stage. Nothing is invented, though.
    Every key, every nesting level and every value type survives, and each cut is marked.
    """
    def line(text):
        return f'<div class="jl">{text}</div>'

    def key(k):
        return f'<span class="jk" id="jkq-{k}">"<span id="jk-{k}">{k}</span>"</span>'

    out = [line("{")]

    title = json.dumps(data["title"])
    out.append(
        '<div class="jg">'
        + line(f'  {key("title")}: <span class="jv" id="jvb-title">{title}</span>,')
        + "</div>"
    )

    records_html = [
        line("    { " + ", ".join(f"{json.dumps(k)}: {json.dumps(v)}" for k, v in row.items()) + " },")
        for row in data["sales"][:records]
    ]
    records_html.append(line("    …"))
    out.append(
        '<div class="jg">'
        + line(f'  {key("sales")}: [')
        + f'<div class="jv" id="jvb-sales">{"".join(records_html)}</div>'
        + line("  ],")
        + "</div>"
    )

    pairs_html = [
        line(f"    {json.dumps(k)}: {json.dumps(v)},")
        for k, v in list(data["hours"].items())[:pairs]
    ]
    pairs_html.append(line("    …"))
    out.append(
        '<div class="jg">'
        + line(f'  {key("hours")}: {{')
        + f'<div class="jv" id="jvb-hours">{"".join(pairs_html)}</div>'
        + line("  }")
        + "</div>"
    )
    out.append(line("}"))
    return "".join(out)


def dense_grid(path, sheet, rows=30, cols=20):
    """The opener: a real government workbook, shown as the wall of data it is."""
    ws = load_workbook(path, data_only=True)[sheet]
    out = []
    for r in range(1, rows + 1):
        cells = [f"<th>{r}</th>"]
        for c in range(1, cols + 1):
            v = ws.cell(row=r, column=c).value
            v = "" if v is None else (f"{v:,.0f}" if isinstance(v, (int, float)) else str(v)[:22])
            cells.append(f"<td>{v}</td>")
        out.append("<tr>" + "".join(cells) + "</tr>")
    letters = "".join(f"<th>{chr(ord('A')+i)}</th>" for i in range(cols))
    return f'<table class="sheet dense"><tr><th></th>{letters}</tr>{"".join(out)}</table>'


def build(show_timeline=False):
    book, spec = HERE / "quarter.xlsx", HERE / "quarter.toml"
    t = timeline()
    data = read_spec(str(book), str(spec))
    rows = tomllib.loads(spec.read_text())["extract"]
    wb = load_workbook(book)
    labels = {k: reference(v) for k, v in rows.items()}
    ranges = {k: resolve(labels[k]["text"], wb) for k in rows}
    spec_rows = "".join(
        f'<div class="srow" id="s{k}" style="--c:{COLOURS[k]}">'
        f'<span class="k">{k}</span> = <span class="v">"{v}"</span></div>'
        for k, v in rows.items()
    )
    with as_file(DATA / "slgfs.xlsx") as slgfs:
        dense = dense_grid(slgfs, "Scotland")
        tabs = load_workbook(slgfs, read_only=True).sheetnames[:26]
    payload = {
        "dense": dense,
        "denseTabs": tabs,
        "grid": grid(book),
        "specRows": spec_rows,
        "json": json_html(data),
        "ranges": ranges,
        "labels": labels,
        "colours": COLOURS,
        "T": t.values(),
        "captions": captions(t),
    }
    html = TEMPLATE.replace("__PAYLOAD__", json.dumps(payload))
    (HERE / "introducer.html").write_text(html, encoding="utf-8")
    print(f"wrote {HERE / 'introducer.html'}")
    if show_timeline:
        print(t.listing())


TEMPLATE = r"""<!doctype html>
<meta charset="utf-8"><title>EDJAS introducer</title>
<style>
  :root { --ink:#1c1a22; --rule:#d8d2e4; --paper:#e8e2f5; }
  * { box-sizing:border-box; }
  html,body { height:100%; overflow:hidden; }
  body { margin:0; background:#3a3348; font-family:system-ui,-apple-system,"Segoe UI",sans-serif;
         display:flex; flex-direction:column; align-items:center; justify-content:center; gap:10px; }
  /* the frame is authored at 1280x720 and shown at whatever fits: #fit carries the scale and
     #sizer carries the scaled footprint, because a transform does not change a layout box */
  #sizer { position:relative; flex:0 0 auto; }
  #fit { position:absolute; left:0; top:0; width:1280px; height:720px; transform-origin:0 0; }
  #stage { position:relative; width:1280px; height:720px; background:#fff; overflow:hidden;
           color:var(--ink); box-shadow:0 12px 60px rgba(0,0,0,.5); }
  .layer { position:absolute; inset:0; }
  /* the clean workbook now sits to the RIGHT of the specification, so a key on the left and
     the values it selects on the right read in the same order as "key": value does in JSON */
  table.sheet { border-collapse:collapse; font-size:14px; position:absolute; left:500px; top:172px; }
  table.dense { position:absolute; left:0; top:0; font-size:10px; }
  table.dense td { min-width:62px; max-width:62px; height:21px; padding:1px 5px;
                   overflow:hidden; white-space:nowrap; text-overflow:ellipsis; }
  table.dense th { height:17px; font-size:9px; width:26px; }
  .sheet th { background:#eee9f5; color:#655c78; font-weight:600; font-size:12px;
              border:1px solid var(--rule); width:34px; height:26px; }
  /* Highlighted cells and lifted rows are painted from t by seek(), not by a CSS
     transition.  A transition runs on the wall clock, which makes the frame at a given t
     depend on how long ago the class changed -- the one thing this animation must never
     do, since every check and every captured frame assumes t is the whole story. */
  .sheet td { border:1px solid var(--rule); padding:3px 7px; min-width:58px; height:26px;
              white-space:nowrap; }
  /* a name belongs to the workbook, so it is drawn as a badge attached to its range;
     an address belongs to the specification, so it is drawn as a typed token */
  #labels .rbox { position:absolute; border-radius:3px; outline-offset:1px; }
  #labels .rbox.name { outline:2px dashed var(--c); }
  #labels .rbox.address { outline:2px solid var(--c); }
  #labels .chip { position:absolute; height:19px; display:flex; align-items:center;
                  padding:0 7px; border-radius:5px; font-size:12px; font-weight:700;
                  white-space:nowrap; }
  #labels .chip.name { background:var(--c); color:#fff; }
  #labels .chip.address { background:#fff; color:var(--c); border:1.5px solid var(--c);
                          font-family:ui-monospace,"SF Mono",Menlo,monospace; }
  #tabs { position:absolute; left:20px; top:660px; display:flex; gap:2px; max-width:1240px;
        overflow:hidden; }
  #tabs span { flex:0 0 auto; font-size:11px; padding:3px 9px; background:#eee9f5;
               border:1px solid var(--rule); border-bottom:none; color:#655c78; }
  #tabs span.on { background:#fff; font-weight:700; color:var(--ink); }
  #spec { position:absolute; left:32px; top:196px; width:380px; background:#faf8fd;
          border:1px solid var(--rule); border-radius:10px; padding:16px 18px;
          font-family:ui-monospace,"SF Mono",Menlo,monospace; font-size:15px; }
  #spec .hdr { color:#8a7fa6; margin-bottom:8px; }
  /* no transform here: the flyers are measured against these boxes, and a scaled row would
     make every key start its journey three pixels away from where it appears to be */
  .srow { padding:5px 8px; border-radius:6px; white-space:nowrap; }
  .srow .k { color:var(--c); font-weight:700; }
  .srow .v { color:#3d3550; }
  /* an <svg> is a replaced element: inset:0 alone leaves it at its intrinsic 300x150,
     which silently clips every thread. It needs real dimensions. */
  #threads { position:absolute; inset:0; width:1280px; height:720px; pointer-events:none; }
  #threads path { fill:none; stroke-width:3; stroke-linecap:round;
                  stroke-dasharray:var(--len); stroke-dashoffset:var(--len); }
  #json { position:absolute; left:50%; top:50%; transform:translate(-50%,-50%) scale(1);
          font-family:ui-monospace,Menlo,monospace; font-size:17px; line-height:1.55;
          background:#1e1a2b; color:var(--paper); padding:22px 30px; border-radius:12px;
          white-space:pre; opacity:0; box-shadow:0 20px 60px rgba(0,0,0,.4); }
  #json .jl { min-height:1.55em; }
  #json .jl .jv { display:inline-block; }
  /* the slots the flyers land in: they hold their space while empty, so the card never reflows */
  #json .jk, #json .jv { opacity:0; }
  /* the travelling keys and values -- exact overlays on the originals until they move */
  #morph { pointer-events:none; opacity:0; }
  .fly { position:absolute; transform-origin:50% 50%; }
  .fly .pad { position:absolute; inset:-6px -13px; border-radius:12px; opacity:0; z-index:-1;
              outline:2.5px solid var(--c); box-shadow:0 10px 30px rgba(0,0,0,.20); }
  .fly .txt { position:absolute; inset:0; display:flex; align-items:center;
              font-family:ui-monospace,"SF Mono",Menlo,monospace; font-size:15px; font-weight:700; }
  .fly .q { position:absolute; top:0; bottom:0; display:flex; align-items:center; opacity:0; }
  .fly .q.l { right:100%; margin-right:.08em; }
  .fly .q.r { left:100%; margin-left:.08em; }
  .fly .cv { position:absolute; display:flex; align-items:center; padding:3px 7px;
             font-size:14px; white-space:nowrap; }
  #caption { position:absolute; left:0; right:0; bottom:52px; text-align:center;
             font-size:30px; font-weight:700; color:var(--ink); opacity:0; }
  #fan { position:absolute; inset:0; display:grid; place-items:center; opacity:0; }
  #fan .ring { position:relative; width:100%; height:100%; }
  #fan .dest { position:absolute; left:50%; top:50%; background:#fff; border:2px solid var(--rule);
               border-radius:10px; padding:11px 18px; font-size:18px; font-weight:600;
               color:var(--ink); box-shadow:0 8px 26px rgba(0,0,0,.18); white-space:nowrap; }
  #cta { position:absolute; inset:0; display:grid; place-items:center; opacity:0; }
  #cta code { font-family:ui-monospace,Menlo,monospace; font-size:34px; background:#1e1a2b;
              color:#fff; padding:14px 26px; border-radius:10px; }
  #scrub { display:flex; gap:12px; align-items:center; color:#fff; flex:0 0 auto;
           font-family:system-ui; font-size:13px; }
  #scrub input { flex:1; }
</style>

<div id="sizer"><div id="fit"><div id="stage">
  <div class="layer" id="denseLayer"></div>
  <div class="layer" id="sheetLayer"></div>
  <div class="layer" id="labels"></div>
  <div class="layer"><div id="spec"><div class="hdr">[extract]</div><div id="specRows"></div></div></div>
  <svg class="layer" id="threads"></svg>
  <div class="layer"><div id="json"></div></div>
  <div class="layer" id="morph"></div>
  <div class="layer" id="fan"><div class="ring"></div></div>
  <div class="layer" id="cta"><code>pip install edjas</code></div>
  <div id="caption"></div>
</div></div></div>
<div id="scrub">
  <button id="play">play</button>
  <input id="t" type="range" min="0" max="35" step="0.02" value="0">
  <span id="tv">0.00s</span>
</div>

<script>
const D = __PAYLOAD__;
const stage = document.getElementById('stage');
document.getElementById('sheetLayer').innerHTML = D.grid;
document.getElementById('denseLayer').innerHTML = D.dense +
  '<div id="tabs">' + D.denseTabs.map((n,i)=>
     `<span class="${i===0?'on':''}">${n}</span>`).join('') + '</div>';
document.getElementById('specRows').innerHTML = D.specRows;
document.getElementById('json').innerHTML = D.json;
/* offsets in px from the centre: the JSON sits there and these travel outward from it */
const DESTS=[['Dashboards',-430,-190],['A web page',300,-215],['An API',-470,60],
             ['A database',360,45],['A report',-60,235]];
document.querySelector('#fan .ring').innerHTML =
  DESTS.map(([n])=>`<div class="dest">${n}</div>`).join('');

/* ---- helpers: every visual is a pure function of t, so seek() is deterministic ---- */
const clamp=(x,a=0,b=1)=>Math.max(a,Math.min(b,x));
const ease =x=>x<.5?4*x*x*x:1-Math.pow(-2*x+2,3)/2;
const ph=(t,a,b)=>ease(clamp((t-a)/(b-a)));            // 0..1 across [a,b]
const win=(t,a,b,f=.4)=>Math.min(ph(t,a,a+f), 1-ph(t,b-f,b));  // fade in, hold, fade out
const rgb=h=>[1,3,5].map(i=>parseInt(h.slice(i,i+2),16));
const mix=(a,b,k)=>{const A=rgb(a),B=rgb(b);           // lerp in sRGB, so colours can travel
  return '#'+A.map((v,i)=>Math.round(v+(B[i]-v)*k).toString(16).padStart(2,'0')).join('');};
/* Every geometry below is in frame coordinates -- 1280x720, whatever the frame is shown at.
   getBoundingClientRect reports the scaled pixels, so dividing by SCALE keeps the animation
   independent of the viewport: resizing changes what you see, never where anything goes. */
let SCALE=1;
function fit(force){
  const pad=20, bar=46;
  SCALE = force || Math.min((innerWidth-pad*2)/1280, (innerHeight-pad*2-bar)/720);
  document.getElementById('fit').style.transform=`scale(${SCALE})`;
  const sz=document.getElementById('sizer');
  sz.style.width=(1280*SCALE)+'px'; sz.style.height=(720*SCALE)+'px';
  document.getElementById('scrub').style.width=(1280*SCALE)+'px';
}
window.fit=fit;                         /* frame capture forces 1:1 */
addEventListener('resize',()=>fit());
const box=el=>{const s=stage.getBoundingClientRect(), r=el.getBoundingClientRect();
  return {x:(r.left-s.left)/SCALE, y:(r.top-s.top)/SCALE, w:r.width/SCALE, h:r.height/SCALE,
          cx:(r.left-s.left+r.width/2)/SCALE, cy:(r.top-s.top+r.height/2)/SCALE};};

const KEYS=Object.keys(D.ranges);
const cells=key=>{                        /* the cells this reference actually selects */
  const [tl,br]=D.ranges[key].split(':');
  const c1=tl.charCodeAt(0), r1=+tl.slice(1), c2=br.charCodeAt(0), r2=+br.slice(1);
  const out=[];
  for(let r=r1;r<=r2;r++) for(let c=c1;c<=c2;c++){
    const el=document.getElementById('c'+String.fromCharCode(c)+r); if(el) out.push(el);
  }
  return out;
};

/* What the workbook offers and what the spec asks for, drawn on the sheet itself: two of
   these ranges carry names the workbook already holds, the third is a bare cell address. */
let TAGS=[];
function buildLabels(){
  const L=document.getElementById('labels'); L.innerHTML=''; TAGS=[];
  for(const key of KEYS){
    const [tl,br]=D.ranges[key].split(':');
    const a=box(document.getElementById('c'+tl)), b=box(document.getElementById('c'+br));
    const w=b.x+b.w-a.x, h=b.y+b.h-a.y, k=D.labels[key].kind;
    const g=document.createElement('div');
    g.style.cssText=`--c:${D.colours[key]};position:absolute;inset:0`;
    g.innerHTML=
      `<div class="rbox ${k}" style="left:${a.x}px;top:${a.y}px;width:${w}px;height:${h}px"></div>`+
      `<div class="chip ${k}" style="left:${a.x}px;top:${a.y-22}px">${D.labels[key].text}</div>`;
    L.appendChild(g); TAGS.push({el:g, kind:k});
  }
}

/* thread geometry, measured from the live layout so it survives restyling */
const svg=document.getElementById('threads');
function buildThreads(){
  svg.innerHTML='';
  for(const key of KEYS){
    const row=box(document.getElementById('s'+key));
    const [tl,br]=D.ranges[key].split(':');
    const a=box(document.getElementById('c'+tl)), b=box(document.getElementById('c'+br));
    const x1=row.x+row.w, y1=row.cy;               /* out of the spec's right-hand edge... */
    const x2=a.x,         y2=(a.y+b.y+b.h)/2;      /* ...into the range's left-hand edge   */
    const p=document.createElementNS('http://www.w3.org/2000/svg','path');
    p.setAttribute('d',`M${x1},${y1} C${x1+70},${y1} ${x2-70},${y2} ${x2},${y2}`);
    p.setAttribute('stroke',D.colours[key]); p.id='p'+key;
    svg.appendChild(p);
    p.style.setProperty('--len', p.getTotalLength());
  }
}

/* The morph. Each key and each block of values gets an exact overlay of the original, so the
   handover is invisible; then the originals fade, the overlays grow, and the overlays fly to
   the boxes their own text will occupy in the JSON -- measured, not guessed. */
let FLY=[];
function buildMorph(){
  const morph=document.getElementById('morph');
  morph.innerHTML=''; FLY=[];
  const j=document.getElementById('json'), sp=document.getElementById('spec');
  const saved=[j.style.opacity, sp.style.transform];
  j.style.opacity=1; sp.style.transform='none';    /* measure the settled layout, not t=0 */
  for(const key of KEYS){
    const c=D.colours[key], tint=mix(c,'#ffffff',0.92);
    document.getElementById('jkq-'+key).style.color=mix(c,'#ffffff',0.28);

    const ks=box(document.querySelector('#s'+key+' .k')), kd=box(document.getElementById('jk-'+key));
    const kf=document.createElement('div');
    kf.className='fly';
    kf.style.cssText=`--c:${c};left:${ks.x}px;top:${ks.y}px;width:${ks.w}px;height:${ks.h}px`;
    kf.innerHTML=`<span class="pad" style="background:${tint}"></span>`+
      `<span class="txt"><span class="q l">"</span>${key}<span class="q r">"</span></span>`;
    morph.appendChild(kf);
    FLY.push({el:kf, txt:kf.querySelector('.txt'), pad:kf.querySelector('.pad'),
              slot:document.getElementById('jkq-'+key), key, grow:1.70, hand:T.handKey,
              quotes:[...kf.querySelectorAll('.q')], sx:0, sy:(KEYS.indexOf(key)-1)*26,
              x0:ks.cx, y0:ks.cy, x1:kd.cx, y1:kd.cy, land:kd.w/ks.w,
              from:c, to:mix(c,'#ffffff',0.28)});

    const cs=cells(key).map(el=>({t:el.textContent, r:box(el)}));
    const x=Math.min(...cs.map(o=>o.r.x)), y=Math.min(...cs.map(o=>o.r.y));
    const w=Math.max(...cs.map(o=>o.r.x+o.r.w))-x, h=Math.max(...cs.map(o=>o.r.y+o.r.h))-y;
    const vf=document.createElement('div');
    vf.className='fly';
    vf.style.cssText=`--c:${c};left:${x}px;top:${y}px;width:${w}px;height:${h}px`;
    vf.innerHTML=`<span class="pad" style="background:${tint}"></span>`+cs.map(o=>
      `<span class="cv" style="left:${o.r.x-x}px;top:${o.r.y-y}px;`+
      `width:${o.r.w}px;height:${o.r.h}px">${o.t}</span>`).join('');
    morph.appendChild(vf);
    const vdEl=document.getElementById('jvb-'+key), vd=box(vdEl); vd.el=vdEl;
    FLY.push({el:vf, txt:vf, pad:vf.querySelector('.pad'), slot:vd.el, key, grow:1.14,
              hand:T.handBlock, quotes:[], sx:(KEYS.indexOf(key)-1)*18, sy:0,
              x0:x+w/2, y0:y+h/2, x1:vd.cx, y1:vd.cy, land:Math.min(vd.w/w, vd.h/h),
              from:'#1c1a22', to:'#e8e2f5'});
  }
  j.style.opacity=saved[0]; sp.style.transform=saved[1];
}

/* The timeline is resolved in Python: see timeline() in make_video.py, where every point
   is fixed either absolutely or against one already fixed, so that moving an anchor moves
   everything hanging off it.  What arrives here is the answer rather than the reasoning --
   a flat table of instants and durations, in seconds, and the captions already anchored to
   the beats they belong to. */
const T = D.T, CAPTIONS = D.captions;
window.T = T;                    /* single values can still be nudged from the console */

function seek(t){
  document.getElementById('t').value=t;
  document.getElementById('tv').textContent=t.toFixed(2)+'s';

  /* the crowded opener calms down, and we cut to the clean workbook */
  const clutter=1-ph(t,T.cutAt,T.cutAt+T.cutFor);
  const strip=ph(t,T.stripAt,T.stripTo);           /* the rest of the detail fades away */
  document.getElementById('denseLayer').style.opacity = clutter;
  const sheetIn=(1-clutter)*(1-strip);
  document.getElementById('sheetLayer').style.opacity = sheetIn;

  /* The names are the workbook's own, so they arrive with the workbook.  The address is not:
     it is what the specification types, so it arrives with the line that types it. */
  document.getElementById('labels').style.opacity = sheetIn;
  TAGS.forEach((g,i)=>{ const at=T.liftAt+i*T.liftPer;
    g.el.style.opacity = g.kind==='name' ? ph(t,T.namesAt,T.namesAt+T.namesFor)
                                         : ph(t,at,at+T.chipFor); });

  /* the specification arrives, from the left */
  const specIn=ph(t,T.specAt,T.specAt+T.specFor);
  const spec=document.getElementById('spec');
  spec.style.opacity = specIn*(1-strip);
  spec.style.transform = `translateX(${(specIn-1)*40}px)`;

  /* the three rows lift and take colour, one after another */
  KEYS.forEach((k,i)=>{
    const at=T.liftAt+i*T.liftPer;
    const up=ph(t,at,at+T.liftFor)*(1-ph(t,T.chromeGone-T.liftFor,T.chromeGone));
    const row=document.getElementById('s'+k);
    row.style.background = up ? `rgba(255,255,255,${up})` : '';
    row.style.boxShadow  = up ? `0 ${8*up}px ${26*up}px rgba(0,0,0,${0.22*up})` : '';
    row.style.outline    = up ? `${2*up}px solid ${D.colours[k]}` : '';
  });

  /* threads draw with staggered starts and a shared arrival, then the cells light */
  KEYS.forEach((k,i)=>{
    const p=document.getElementById('p'+k); if(!p) return;
    const len=p.getTotalLength();
    const drawn=ph(t, T.threadAt+i*T.threadPer, T.threadTo);
    p.style.strokeDashoffset = len*(1-drawn);
    p.style.opacity = 1-ph(t,T.morphAt,T.chromeGone);
    /* the cells come up over the last of the thread's draw, and go out with the sheet */
    const lit=clamp((drawn-T.threadLit)/(1-T.threadLit))
              *(1-ph(t,T.litTo-T.litOutFor,T.litTo));
    const [tl,br]=D.ranges[k].split(':');
    const [c1,r1]=[tl[0],+tl.slice(1)], [c2,r2]=[br[0],+br.slice(1)];
    for(let c=c1.charCodeAt(0);c<=c2.charCodeAt(0);c++)
      for(let r=r1;r<=r2;r++){
        const cell=document.getElementById('c'+String.fromCharCode(c)+r); if(!cell) continue;
        cell.style.boxShadow = lit ? `inset 0 0 0 ${3*lit}px ${D.colours[k]}` : '';
        cell.style.background = lit ? mix('#ffffff',D.colours[k],0.12*lit) : '';
      }
  });

  /* the keys and the values grow, everything else goes, and the two become the JSON */
  const grow=ph(t,T.morphAt,T.morphAt+T.growFor);
  document.getElementById('morph').style.opacity = t>=T.showAt ? 1 : 0;
  FLY.forEach(f=>{
    const skew=KEYS.indexOf(f.key)*T.flightPer, land=T.flightTo+skew;
    const m=ph(t, T.flightAt+skew, land);                        /* the flight */
    const fade=ph(t, land-T.handLead, land-T.handLead+f.hand);   /* and the handover */
    const s=(1+(f.grow-1)*grow)+(f.land-f.grow)*m;
    const gx=f.sx*grow, gy=f.sy*grow;                   /* held apart while enlarged... */
    f.el.style.transform=                               /* ...and released into the flight */
      `translate(${gx+(f.x1-f.x0-gx)*m}px,${gy+(f.y1-f.y0-gy)*m}px) scale(${s*(1+0.26*fade)})`;
    f.el.style.opacity = 1-fade;
    f.pad.style.opacity = ph(t,T.morphAt,T.morphAt+T.padFor)*(1-ph(t,T.flightAt,T.padTo));
    f.txt.style.color = mix(f.from, f.to, m);
    f.quotes.forEach(q=>q.style.opacity=m);
    f.slot.style.opacity = ph(t, land-T.slotLead, land-T.slotLead+f.hand);
  });
  const j=document.getElementById('json');
  const shrink=ph(t,T.shrinkAt,T.shrinkTo);              /* makes room for the destinations */
  const appear=ph(t,T.cardAt,T.cardTo);
  j.style.opacity = appear;
  j.style.transform = `translate(-50%,-50%) translateY(${-46*shrink}px) `+
                      `scale(${(0.94+0.06*appear)*(1-0.38*shrink)})`;

  /* where it can go next */
  document.getElementById('fan').style.opacity = ph(t,T.fanAt,T.fanAt+T.fanFor);
  [...document.querySelectorAll('#fan .dest')].forEach((d,i)=>{
    const at=T.destAt+i*T.destPer;
    const k=ph(t,at,at+T.destFor);                       /* travel out from the JSON */
    const [,dx,dy]=DESTS[i];
    d.style.transform=`translate(-50%,-50%) translate(${dx*k}px,${dy*k}px) scale(${0.6+0.4*k})`;
    d.style.opacity=k;
  });

  const [text,from,to,fade]=CAPTIONS.find(c=>t<c[2]) || CAPTIONS[CAPTIONS.length-1];
  const cap=document.getElementById('caption');
  cap.textContent=text; cap.style.opacity=win(t,from,to,fade);

  document.getElementById('cta').style.opacity = ph(t,T.ctaAt,T.ctaTo);
}
window.seek=seek;                       /* exposed for headless frame capture */

/* preview transport */
let playing=false, t0=0, base=0;
function frame(ts){ if(!playing) return;
  const t=base+(ts-t0)/1000; if(t>=T.end){ seek(T.end); playing=false; return; }
  seek(t); requestAnimationFrame(frame); }
document.getElementById('play').onclick=()=>{
  playing=!playing; base=+document.getElementById('t').value;
  if(playing) requestAnimationFrame(ts=>{t0=ts; frame(ts);}); };
document.getElementById('t').oninput=e=>{ playing=false; seek(+e.target.value); };

document.getElementById('t').max=T.end;
addEventListener('load',()=>{ fit(); buildLabels(); buildThreads(); buildMorph(); seek(0); });
</script>
"""

if __name__ == "__main__":
    build(show_timeline="--timeline" in sys.argv)
