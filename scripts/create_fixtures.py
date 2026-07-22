"""Regenerate the test fixtures used by the EDJAS test suite.

Writes a **data-only** workbook (``tests/data/parameters.xlsx``) — the spreadsheet
EDJAS reads but never modifies — together with a companion extraction spec
(``tests/data/parameters.toml``). The workbook carries no EDJAS markup; all
extraction instructions live in the spec.

Run from the project root:  ``python scripts/create_fixtures.py``
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

DATA = Path(__file__).resolve().parent.parent / "tests" / "data"
XLSX = DATA / "parameters.xlsx"
TOML = DATA / "parameters.toml"

HOURS = [
    ("Monday", "7:00 am - 8:00 pm"),
    ("Tuesday", "7:00 am - 8:00 pm"),
    ("Wednesday", "7:00 am - 8:00 pm"),
    ("Thursday", "7:00 am - 8:00 pm"),
    ("Friday", "7:00 am - 8:00 pm"),
    ("Saturday", "9:00 am - 5:00 pm"),
    ("Sunday", "Closed"),
]
PRICES = [("Tea", 3.25), ("Coffee", 4.0), ("Bacon Sandwich", 8.25)]
H_VECTOR = ["the", "quick", "brown", "fox"]
V_VECTOR = ["jumps", "over", "the", "lazy", "dog"]

SPEC = """\
# EDJAS extraction spec for parameters.xlsx
[extract]
title = "Title"
hours = "{Hours}"
prices = "{Prices}"
hours_list = "[Hours]"
h_vector = "[HVector]"
v_vector = "[VVector]"
"""


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws["A1"] = "EDJAS Demo"  # scalar title

    for offset, (day, opening) in enumerate(HOURS):  # Hours table -> C1:D7
        ws.cell(row=1 + offset, column=3, value=day)
        ws.cell(row=1 + offset, column=4, value=opening)

    for offset, (item, price) in enumerate(PRICES):  # Prices table -> F1:G3
        ws.cell(row=1 + offset, column=6, value=item)
        ws.cell(row=1 + offset, column=7, value=price)

    for offset, word in enumerate(H_VECTOR):  # horizontal vector -> I1:L1
        ws.cell(row=1, column=9 + offset, value=word)

    for offset, word in enumerate(V_VECTOR):  # vertical vector -> I3:I7
        ws.cell(row=3 + offset, column=9, value=word)

    wb.defined_names.add(DefinedName("Title", attr_text="Data!$A$1"))
    wb.defined_names.add(DefinedName("Hours", attr_text="Data!$C$1:$D$7"))
    wb.defined_names.add(DefinedName("Prices", attr_text="Data!$F$1:$G$3"))
    wb.defined_names.add(DefinedName("HVector", attr_text="Data!$I$1:$L$1"))
    wb.defined_names.add(DefinedName("VVector", attr_text="Data!$I$3:$I$7"))

    DATA.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    print(f"Wrote workbook: {XLSX}")


def build_spec():
    TOML.write_text(SPEC)
    print(f"Wrote spec: {TOML}")


if __name__ == "__main__":
    build_workbook()
    build_spec()
